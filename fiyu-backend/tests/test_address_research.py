from __future__ import annotations

import csv
import json
import sqlite3
from copy import deepcopy
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace

import pytest
from pydantic import BaseModel

from fiyu.address_geocoder import (
    AddressGeocodeResult,
    DigitalAgencyAbrGeocoder,
    JsonFileAddressGeocoder,
)
from fiyu.address_geocoding import (
    export_geocoding_inputs,
    geocode_address_file,
    geocode_verified_addresses,
    validate_geocode,
)
from fiyu.address_research import (
    ADDRESS_RESEARCH_INSTRUCTIONS,
    AddressCallMetadata,
    AddressResearchCall,
    AddressResearchResult,
    AddressSourceEvidence,
    ConflictingAddressCandidate,
    address_resolution_status,
    combined_address_call,
    compare_address_components,
    evaluate_address_result,
    generate_address_queries,
    persist_address_call,
    recalculate_address_decisions,
    record_generated_queries,
    run_address_discovery,
    start_address_run,
)
from fiyu.address_review import export_address_review, import_address_review
from fiyu.database import SCHEMA, connect
from fiyu.location_corrections import replace_location
from fiyu.public_catalog import ensure_public_schema
from fiyu.public_score import InternalSignals, calculate_fiyu_score
from fiyu.research_worker import RestaurantResearch, run_research_batch


def _sqlite_artifact_state(path: Path) -> dict[str, bytes | None]:
    return {
        suffix: artifact.read_bytes() if artifact.is_file() else None
        for suffix in ("", "-wal", "-shm", "-journal")
        for artifact in (Path(f"{path}{suffix}"),)
    }


def _sqlite_backup_without_sidecars(source_path: Path, target_path: Path) -> Path:
    source = sqlite3.connect(source_path)
    target = sqlite3.connect(target_path)
    try:
        source.backup(target)
    finally:
        target.close()
        source.close()
    assert not Path(f"{target_path}-wal").exists()
    assert not Path(f"{target_path}-shm").exists()
    return target_path


class FakeResponses:
    def __init__(self, outputs):
        self.outputs = iter(outputs)
        self.calls: list[dict[str, object]] = []

    def _next(self, kwargs):
        self.calls.append(kwargs)
        output = next(self.outputs)
        if isinstance(output, BaseException):
            raise output
        return output

    def parse(self, **kwargs):
        return self._next(kwargs)

    def create(self, **kwargs):
        return self._next(kwargs)


class FakeClient:
    def __init__(self, outputs):
        self.responses = FakeResponses(outputs)


def _candidate(place_id="p1", **changes):
    value = {
        "place_id": place_id,
        "name_ja": "鮨さいとう",
        "name_en": "Sushi Saito",
        "category": "sushi",
        "discovery_area": "Taito",
        "discovery_area_type": "ward",
        "discovery_areas_json": json.dumps(
            [{"area": "Taito", "area_type": "ward", "source_file": "Taito.csv"}]
        ),
    }
    value.update(changes)
    return value


def _source(source_type="official_restaurant_website", **changes):
    value = {
        "source_type": source_type,
        "source_url": "https://sushi-saito.example.jp/access",
        "source_title": "鮨さいとう 店舗情報",
        "source_language": "ja",
        "accessed_at": "2026-07-27",
        "address_text_as_displayed": "東京都台東区谷中1丁目2-3",
        "identity_evidence_summary": "The page identifies 鮨さいとう.",
        "address_evidence_summary": "The access page displays the street address.",
        "restaurant_controlled": True,
        "supports_candidate_address": True,
        "warnings": [],
    }
    value.update(changes)
    return AddressSourceEvidence(**value)


def _result(**changes):
    value = {
        "identity_status": "confirmed",
        "identity_confidence": 0.97,
        "matched_name": "鮨さいとう",
        "branch_name": None,
        "address_raw": "東京都台東区谷中1丁目2-3",
        "postal_code": "110-0001",
        "prefecture": "東京都",
        "municipality_or_ward": "台東区",
        "neighborhood": "谷中",
        "street_or_block": "1丁目2-3",
        "building": None,
        "source_evidence": [_source()],
        "conflicting_address_candidates": [],
        "search_queries_attempted": ['"鮨さいとう" 台東区 住所'],
        "warnings": [],
        "recommended_action": "accept_if_deterministic_rules_pass",
        "research_summary": "One restaurant-controlled source displays the address.",
    }
    value.update(changes)
    return AddressResearchResult(**value)


def _response(
    parsed,
    *,
    search_status="completed",
    queries=None,
    response_status="completed",
    incomplete_reason=None,
    raw_output=None,
    usage_values=None,
    include_web_search=True,
    response_id="resp_1",
):
    queries = queries or ['"鮨さいとう" Taito 住所']
    web_call = SimpleNamespace(
        type="web_search_call",
        id="ws_1",
        status=search_status,
        action=SimpleNamespace(type="search", queries=queries, query=None, sources=[]),
    )
    citation = SimpleNamespace(
        type="url_citation",
        url="https://citation.example.jp/shop",
        title="Citation title",
    )
    if raw_output is None:
        if isinstance(parsed, BaseModel):
            raw_output = parsed.model_dump_json()
        elif parsed is not None:
            raw_output = json.dumps(parsed, ensure_ascii=False)
        else:
            raw_output = ""
    message = SimpleNamespace(
        type="message",
        content=[
            SimpleNamespace(
                type="output_text", text=raw_output, annotations=[citation]
            )
        ],
    )
    usage_values = usage_values or {}
    usage = SimpleNamespace(
        input_tokens=usage_values.get("input_tokens", 100),
        input_tokens_details=SimpleNamespace(
            cached_tokens=usage_values.get("cached_input_tokens", 25), cache_write_tokens=0
        ),
        output_tokens=usage_values.get("output_tokens", 40),
        output_tokens_details=SimpleNamespace(
            reasoning_tokens=usage_values.get("reasoning_tokens", 10)
        ),
        total_tokens=usage_values.get("total_tokens", 140),
    )
    return SimpleNamespace(
        output_parsed=parsed,
        output_text=raw_output,
        id=response_id,
        model="configured-model",
        status=response_status,
        incomplete_details=(
            SimpleNamespace(reason=incomplete_reason) if incomplete_reason else None
        ),
        error=None,
        output=([web_call] if include_web_search else []) + [message],
        usage=usage,
    )


def _db(tmp_path):
    path = tmp_path / "address.db"
    ensure_public_schema(path)
    with connect(path) as connection:
        for item in (
            _candidate(),
            _candidate(
                "ambiguous",
                name_ja="金寿司",
                location_resolution_reason="unresolved_ambiguous_exact_name_candidates",
            ),
            _candidate(
                "manual",
                name_ja="楽亀",
                location_resolution_reason="needs_manual_review",
            ),
            _candidate("unpublished", is_published=0),
            _candidate(
                "ChIJ2WzWhfWPGGARyYQS7SD2tIM",
                name_ja="金すし",
                map_display_eligible=1,
                location_resolution_reason="auto_verified_exact_name_with_geographic_corroboration",
            ),
        ):
            connection.execute(
                """
                INSERT INTO public_restaurants (
                    place_id, name_ja, name_en, primary_category, discovery_area,
                    discovery_area_type, discovery_areas_json, location_resolution_reason,
                    location_verification_status, map_display_eligible, address_resolution_status,
                    research_status, is_published, fiyu_score, fiyu_confidence, why_fiyu,
                    food_tags_json, signature_dishes_json, evidence_json, evidence_urls_json,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'address_not_researched',
                          'complete', ?, 80, 90, 'Kept description.', '["sushi"]',
                          '["おまかせ"]', '{"kept":true}', '[]', 'created', 'before')
                """,
                (
                    item["place_id"],
                    item["name_ja"],
                    item["name_en"],
                    item["category"],
                    item["discovery_area"],
                    item["discovery_area_type"],
                    item["discovery_areas_json"],
                    item.get("location_resolution_reason", "likely_not_represented_in_osm"),
                    item.get("location_verification_status", "unresolved"),
                    item.get("map_display_eligible", 0),
                    item.get("is_published", 1),
                ),
            )
        connection.commit()
    return path


def _stored(path, place_id="p1"):
    with connect(path) as connection:
        return dict(
            connection.execute(
                "SELECT * FROM public_restaurants WHERE place_id=?", (place_id,)
            ).fetchone()
        )


def _call(candidate, result):
    metadata = AddressCallMetadata(
        response_id="resp_1",
        model="configured-model",
        search_calls=(
            {
                "id": "ws_1",
                "status": "completed",
                "action_type": "search",
                "queries": result.search_queries_attempted,
                "sources": [],
            },
        ),
        response_request_count=1,
        web_search_action_count=1,
        input_tokens=100,
        cached_input_tokens=25,
        output_tokens=40,
        reasoning_tokens=10,
        total_tokens=140,
    )
    return AddressResearchCall(
        result=result,
        acceptance=evaluate_address_result(candidate, result),
        generated_queries=tuple(result.search_queries_attempted),
        metadata=metadata,
    )


def _persist(path, result, place_id="p1"):
    candidate = _candidate(place_id)
    run_id = start_address_run(
        path,
        place_id=place_id,
        model="configured-model",
        forced=False,
        combined_research=False,
    )
    record_generated_queries(
        path,
        run_id=run_id,
        place_id=place_id,
        queries=result.search_queries_attempted,
    )
    evidence_id = persist_address_call(
        path,
        place_id=place_id,
        run_id=run_id,
        call=_call(candidate, result),
    )
    return evidence_id


def _persist_for_candidate(path, candidate, result):
    place_id = candidate["place_id"]
    run_id = start_address_run(
        path,
        place_id=place_id,
        model="configured-model",
        forced=False,
        combined_research=False,
    )
    return persist_address_call(
        path,
        place_id=place_id,
        run_id=run_id,
        call=_call(candidate, result),
    )


@pytest.mark.parametrize(
    ("source_type", "url", "controlled"),
    [
        ("official_restaurant_website", "https://shop.example.jp/access", True),
        ("official_restaurant_social_profile", "https://facebook.com/shop/about", True),
        ("restaurant_controlled_reservation_page", "https://reserve.example.jp/shop", True),
        ("local_government_or_official_listing", "https://city.taito.lg.jp/shop", False),
    ],
)
def test_single_strong_explicit_source_is_accepted(source_type, url, controlled):
    result = _result(
        source_evidence=[
            _source(source_type, source_url=url, restaurant_controlled=controlled)
        ]
    )
    acceptance = evaluate_address_result(_candidate(), result)
    assert acceptance.status == "accepted"
    assert acceptance.resolution_status == "address_verified"


def test_two_independent_secondary_sources_corroborate_address():
    result = _result(
        source_evidence=[
            _source(
                "permitted_business_directory",
                source_url="https://directory.example.jp/saito",
                restaurant_controlled=False,
            ),
            _source(
                "established_local_editorial_source",
                source_url="https://local-news.example.com/saito",
                restaurant_controlled=False,
            ),
        ]
    )
    acceptance = evaluate_address_result(_candidate(), result)
    assert acceptance.status == "provisional"
    assert acceptance.confidence_tier == "provisional_high"


def test_snippet_alone_and_restricted_sources_are_explicitly_provisional():
    snippet = _result(
        source_evidence=[
            _source(
                "search_result_snippet",
                source_url="https://search.example/result",
                restaurant_controlled=False,
            )
        ]
    )
    restricted = _result(
        source_evidence=[
            _source(
                "official_restaurant_social_profile",
                source_url="https://instagram.com/sushi_saito",
                restaurant_controlled=True,
            )
        ]
    )
    assert evaluate_address_result(_candidate(), snippet).confidence_tier == "provisional_medium"
    assert evaluate_address_result(_candidate(), restricted).confidence_tier == "provisional_medium"


def test_restricted_source_requires_exact_restaurant_and_branch_identity():
    candidate = _candidate(
        name_ja="鮨さいとう 六本木店",
        title="鮨さいとう 六本木店",
    )
    exact_branch = _result(
        matched_name="鮨さいとう 六本木店",
        branch_name="六本木店",
        source_evidence=[
            _source(
                "lead_only_restricted_platform",
                source_url="https://tabelog.com/tokyo/example",
                restaurant_controlled=False,
            )
        ],
    )
    accepted = evaluate_address_result(candidate, exact_branch)
    assert accepted.status == "provisional"
    assert accepted.confidence_tier == "provisional_medium"

    fuzzy_only = exact_branch.model_copy(update={"matched_name": "鮨さいとう 六本木"})
    blocked = evaluate_address_result(candidate, fuzzy_only)
    assert blocked.status == "needs_review"
    assert "restricted_source_requires_exact_identity" in blocked.reasons


def test_one_permitted_directory_source_becomes_provisional_high():
    result = _result(
        source_evidence=[
            _source(
                "permitted_business_directory",
                source_url="https://directory.example.jp/saito",
                restaurant_controlled=False,
            )
        ]
    )
    acceptance = evaluate_address_result(_candidate(), result)
    assert acceptance.status == "provisional"
    assert acceptance.resolution_status == "address_provisionally_accepted"
    assert acceptance.confidence_tier == "provisional_high"


def test_two_independent_weak_sources_become_provisional_high():
    result = _result(
        source_evidence=[
            _source(
                "search_result_snippet", source_url="https://lead-one.example/saito",
                restaurant_controlled=False,
            ),
            _source(
                "weak_user_generated_content", source_url="https://lead-two.example/saito",
                restaurant_controlled=False,
            ),
        ]
    )
    acceptance = evaluate_address_result(_candidate(), result)
    assert acceptance.status == "provisional"
    assert acceptance.confidence_tier == "provisional_high"


def test_probable_identity_with_consistent_core_becomes_provisional_medium():
    result = _result(identity_status="probable", identity_confidence=0.78)
    acceptance = evaluate_address_result(_candidate(), result)
    assert acceptance.status == "provisional"
    assert acceptance.confidence_tier == "provisional_medium"
    assert "identity_probable_but_consistent" in acceptance.reasons


def test_wrong_ward_branch_ambiguity_and_non_street_address_require_review():
    wrong_ward = _result(
        address_raw="東京都杉並区高円寺1丁目2-3",
        municipality_or_ward="杉並区",
    )
    branch = _result(branch_name="六本木店")
    ward_only = _result(address_raw="東京都台東区谷中", street_or_block=None)
    assert "material_component_conflict:municipality_or_ward" in evaluate_address_result(
        _candidate(), wrong_ward
    ).reasons
    assert "unresolved_branch_ambiguity" in evaluate_address_result(
        _candidate(), branch
    ).reasons
    assert "address_is_not_explicit_street_level" in evaluate_address_result(
        _candidate(), ward_only
    ).reasons


def test_conflicting_addresses_and_not_found_are_separate_states():
    conflicting = _result(
        identity_status="conflicting",
        conflicting_address_candidates=[
            ConflictingAddressCandidate(
                address_raw="東京都中央区銀座1-1-1",
                source_urls=["https://other.example"],
                summary="A different branch address.",
            )
        ],
    )
    not_found = _result(
        identity_status="not_found",
        identity_confidence=0,
        matched_name=None,
        address_raw=None,
        source_evidence=[],
    )
    assert evaluate_address_result(_candidate(), conflicting).status == "conflicting"
    assert evaluate_address_result(_candidate(), not_found).status == "not_found"


def _chiyoda_candidate():
    return _candidate(
        name_ja="牛たんの檸檬 秋葉原店",
        name_en="Gyutan no Lemon Akihabara",
        discovery_area="Chiyoda",
        discovery_areas_json=json.dumps([{"area": "Chiyoda", "area_type": "ward"}]),
    )


def _building_conflict_result(*, second_street="3-38", second_ward="千代田区"):
    core = "東京都千代田区神田佐久間町3-38"
    return AddressResearchResult(
        identity_status="confirmed",
        identity_confidence=0.98,
        matched_name="牛たんの檸檬 秋葉原店",
        address_raw=f"{core} 大陽ビル1階",
        prefecture="東京都",
        municipality_or_ward="千代田区",
        neighborhood="神田佐久間町",
        street_or_block="3-38",
        building="大陽ビル",
        floor="1階",
        source_evidence=[
            AddressSourceEvidence(
                source_type="official_restaurant_website",
                source_url="https://gyutan.example.jp/akihabara",
                source_language="ja",
                accessed_at="2026-07-27",
                address_text_as_displayed=f"{core} 大陽ビル1階",
                identity_evidence_summary="Exact branch name appears on the access page.",
                address_evidence_summary="The page supplies the full address.",
                restaurant_controlled=True,
                supports_candidate_address=True,
            ),
            AddressSourceEvidence(
                source_type="restaurant_controlled_reservation_page",
                source_url="https://tablecheck.example/akihabara",
                source_language="ja",
                accessed_at="2026-07-27",
                address_text_as_displayed=(
                    f"東京都{second_ward}神田佐久間町{second_street} 第5東ビル B1F"
                ),
                street_or_block=second_street,
                municipality_or_ward=second_ward,
                building="第5東ビル",
                floor="B1F",
                identity_evidence_summary="Exact branch name appears on the reservation page.",
                address_evidence_summary="The reservation page supplies an address.",
                restaurant_controlled=True,
                supports_candidate_address=False,
            ),
        ],
        conflicting_address_candidates=[
            ConflictingAddressCandidate(
                address_raw=f"東京都{second_ward}神田佐久間町{second_street} 第5東ビル B1F",
                source_urls=["https://tablecheck.example/akihabara"],
                street_or_block=second_street,
                municipality_or_ward=second_ward,
                building="第5東ビル",
                floor="B1F",
            )
        ],
    )


def test_building_and_floor_conflicts_verify_only_the_agreed_core():
    result = _building_conflict_result()
    agreement = compare_address_components(result)
    assert agreement.agreed_core_address == "東京都千代田区神田佐久間町3-38"
    assert agreement.core_address_verified is True
    assert agreement.full_address_verified is False
    assert set(agreement.non_material_conflicting_components) >= {"building", "floor"}
    assert agreement.material_conflicting_components == []
    assert evaluate_address_result(_chiyoda_candidate(), result).status == "accepted"


def test_floor_only_conflict_is_non_material():
    result = _building_conflict_result()
    second = result.source_evidence[1].model_copy(
        update={
            "address_text_as_displayed": "東京都千代田区神田佐久間町3-38 大陽ビル B1F",
            "building": "大陽ビル",
        }
    )
    conflict = result.conflicting_address_candidates[0].model_copy(
        update={
            "address_raw": "東京都千代田区神田佐久間町3-38 大陽ビル B1F",
            "building": "大陽ビル",
        }
    )
    result = result.model_copy(
        update={"source_evidence": [result.source_evidence[0], second],
                "conflicting_address_candidates": [conflict]}
    )
    agreement = compare_address_components(result)
    assert agreement.non_material_conflicting_components == ["floor"]
    assert agreement.core_address_verified and not agreement.full_address_verified


@pytest.mark.parametrize(
    ("changes", "component"),
    [
        ({"second_street": "3-39"}, "street_or_block"),
        ({"second_ward": "台東区"}, "municipality_or_ward"),
    ],
)
def test_street_and_ward_conflicts_remain_material(changes, component):
    result = _building_conflict_result(**changes)
    agreement = compare_address_components(result)
    assert component in agreement.material_conflicting_components
    assert agreement.core_address_verified is False
    assert evaluate_address_result(_chiyoda_candidate(), result).status == "conflicting"


def test_japanese_street_notation_and_full_width_digits_compare_equally():
    result = _building_conflict_result()
    second = result.source_evidence[1].model_copy(
        update={
            "address_text_as_displayed": "東京都千代田区神田佐久間町１丁目１３",
            "street_or_block": "１丁目１３",
            "building": None,
            "floor": None,
        }
    )
    result = result.model_copy(
        update={
            "address_raw": "東京都千代田区神田佐久間町1-13",
            "street_or_block": "1-13",
            "building": None,
            "floor": None,
            "source_evidence": [result.source_evidence[0].model_copy(update={
                "address_text_as_displayed": "東京都千代田区神田佐久間町1-13",
                "street_or_block": "1-13",
                "building": None,
                "floor": None,
            }), second],
            "conflicting_address_candidates": [],
        }
    )
    agreement = compare_address_components(result)
    assert agreement.street_or_block_agreement == "agrees"
    assert agreement.agreed_core_address == "東京都千代田区神田佐久間町1-13"


@pytest.mark.parametrize("variant", ["3‐30‐5", "3－30－5", "３ー３０ー５"])
def test_japanese_hyphen_variants_compare_equally(variant):
    result = _result(street_or_block="3-30-5", address_raw="東京都台東区谷中3-30-5")
    source = result.source_evidence[0].model_copy(
        update={"street_or_block": variant, "address_text_as_displayed": f"東京都台東区谷中{variant}"}
    )
    agreement = compare_address_components(result.model_copy(update={"source_evidence": [source]}))
    assert agreement.street_or_block_agreement == "agrees"


@pytest.mark.parametrize(
    "variant",
    ["2丁目27−16", "2丁目27-16", "2-chōme−27−１６", "2 Chome 27-16", "２丁目２７−１６"],
)
def test_japanese_and_romanized_chome_notation_compare_structurally(variant):
    result = _result(
        street_or_block="2-27-16",
        address_raw="東京都台東区浅草2-27-16",
    )
    source = result.source_evidence[0].model_copy(
        update={
            "street_or_block": variant,
            "address_text_as_displayed": f"東京都台東区浅草{variant}",
        }
    )
    agreement = compare_address_components(
        result.model_copy(update={"source_evidence": [source]})
    )
    assert agreement.street_or_block_agreement == "agrees"
    assert agreement.material_conflicting_components == []


@pytest.mark.parametrize("variant", ["2丁目28-16", "3丁目27-16", "2丁目27-18"])
def test_structurally_different_japanese_address_numbers_still_conflict(variant):
    result = _result(
        street_or_block="2-27-16",
        address_raw="東京都台東区浅草2-27-16",
    )
    source = result.source_evidence[0].model_copy(
        update={
            "street_or_block": variant,
            "address_text_as_displayed": f"東京都台東区浅草{variant}",
        }
    )
    agreement = compare_address_components(
        result.model_copy(update={"source_evidence": [source]})
    )
    assert agreement.street_or_block_agreement == "conflicts"
    assert agreement.material_conflicting_components == ["street_or_block"]


def test_cross_language_tokyo_kojimachi_components_and_split_chome_agree():
    result = _result(
        prefecture="東京都",
        municipality_or_ward="千代田区",
        neighborhood="麹町1丁目",
        street_or_block="6-30",
        floor="1階",
        address_raw="東京都千代田区麹町1丁目6-30 1階",
    )
    source = result.source_evidence[0].model_copy(
        update={
            "prefecture": "Tokyo Metropolis",
            "municipality_or_ward": "Chiyoda Ward",
            "neighborhood": "Kojimachi 1 Chome",
            "street_or_block": "6-30",
            "floor": "1F",
            "address_text_as_displayed": (
                "Kojimachi 1 Chome 6-30, Chiyoda Ward, Tokyo Metropolis 1F"
            ),
        }
    )
    second = source.model_copy(
        update={
            "neighborhood": "麹町",
            "street_or_block": "1-6-30",
            "address_text_as_displayed": "東京都千代田区麹町1-6-30 1F",
        }
    )
    agreement = compare_address_components(
        result.model_copy(update={"source_evidence": [source, second]})
    )
    assert agreement.prefecture_agreement == "agrees"
    assert agreement.municipality_or_ward_agreement == "agrees"
    assert agreement.neighborhood_agreement == "agrees"
    assert agreement.street_or_block_agreement == "agrees"
    assert agreement.floor_agreement == "agrees"


def test_navigation_reference_is_preserved_but_excluded_from_conflict_voting():
    navigation = ConflictingAddressCandidate(
        address_raw="東京都千代田区麹町1-6-4 隣のビルになります。",
        municipality_or_ward="千代田区",
        neighborhood="麹町",
        street_or_block="1-6-4",
        summary="Neighboring-building navigation reference, not the restaurant address.",
    )
    result = _result(
        municipality_or_ward="千代田区",
        neighborhood="麹町",
        street_or_block="1-6-30",
        address_raw="東京都千代田区麹町1-6-30",
        source_evidence=[],
        conflicting_address_candidates=[navigation],
    )
    agreement = compare_address_components(result)
    assert agreement.street_or_block_agreement != "conflicts"
    assert agreement.excluded_non_address_evidence[0]["address_text"] == navigation.address_raw
    assert result.conflicting_address_candidates[0] == navigation


def test_real_alternate_restaurant_address_still_votes_as_a_conflict():
    alternate = ConflictingAddressCandidate(
        address_raw="東京都千代田区麹町1-6-31",
        municipality_or_ward="千代田区",
        neighborhood="麹町",
        street_or_block="1-6-31",
        summary="Alternate current restaurant address.",
    )
    result = _result(
        municipality_or_ward="千代田区",
        neighborhood="麹町",
        street_or_block="1-6-30",
        address_raw="東京都千代田区麹町1-6-30",
        source_evidence=[],
        conflicting_address_candidates=[alternate],
    )
    assert compare_address_components(result).street_or_block_agreement == "conflicts"


@pytest.mark.parametrize(
    ("first", "second"),
    [("1階", "1F"), ("地下1階", "B1F")],
)
def test_floor_notation_equivalents_compare_equally(first, second):
    result = _result(floor=first)
    source = result.source_evidence[0].model_copy(update={"floor": second})
    agreement = compare_address_components(result.model_copy(update={"source_evidence": [source]}))
    assert agreement.floor_agreement == "agrees"


def test_genuinely_different_street_numbers_are_not_equivalent():
    result = _result(street_or_block="3-30-5", address_raw="東京都台東区谷中3-30-5")
    source = result.source_evidence[0].model_copy(
        update={"street_or_block": "4-14-7", "address_text_as_displayed": "東京都台東区谷中4-14-7"}
    )
    agreement = compare_address_components(result.model_copy(update={"source_evidence": [source]}))
    assert agreement.street_or_block_agreement == "conflicts"


def test_explicit_historical_address_is_retained_but_excluded_from_active_conflicts():
    current_raw = "東京都杉並区浜田山3-30-5"
    historical_raw = "東京都杉並区浜田山4-14-7"
    result = AddressResearchResult(
        identity_status="probable",
        identity_confidence=0.93,
        matched_name="浜田山叙々苑",
        address_raw=current_raw,
        prefecture="東京都",
        municipality_or_ward="杉並区",
        neighborhood="浜田山",
        street_or_block="3-30-5",
        source_evidence=[
            _source(
                "permitted_business_directory",
                source_url="https://directory.example/current",
                address_text_as_displayed=current_raw,
                municipality_or_ward="杉並区",
                neighborhood="浜田山",
                street_or_block="3-30-5",
                restaurant_controlled=False,
                address_evidence_summary="現住所を表示。",
            ),
            _source(
                "permitted_business_directory",
                source_url="https://directory.example/old",
                address_text_as_displayed=historical_raw,
                municipality_or_ward="杉並区",
                neighborhood="浜田山",
                street_or_block="4-14-7",
                restaurant_controlled=False,
                address_evidence_summary="旧住所として表示。",
            ),
        ],
        conflicting_address_candidates=[
            ConflictingAddressCandidate(
                address_raw=historical_raw,
                street_or_block="4-14-7",
                summary="移転前の住所。",
            )
        ],
    )
    agreement = compare_address_components(result)
    assert agreement.street_or_block_agreement == "agrees"
    assert agreement.material_conflicting_components == []
    assert {item["address_text"] for item in agreement.excluded_temporal_evidence} == {
        historical_raw
    }
    assert result.source_evidence[1].address_text_as_displayed == historical_raw


def test_address_followed_by_move_language_is_historical_but_unknown_difference_conflicts():
    old = _source(
        address_text_as_displayed="このお店は「東京都台東区谷中4-14-7」から移転しています。",
        street_or_block="4-14-7",
        address_evidence_summary="住所掲載。",
    )
    current = _result(
        street_or_block="3-30-5",
        address_raw="東京都台東区谷中3-30-5",
        source_evidence=[old],
    )
    assert compare_address_components(current).material_conflicting_components == []
    unknown = old.model_copy(
        update={
            "address_text_as_displayed": "東京都台東区谷中4-14-7",
            "address_evidence_summary": "住所掲載。",
        }
    )
    assert compare_address_components(
        current.model_copy(update={"source_evidence": [unknown]})
    ).material_conflicting_components == ["street_or_block"]


def test_page_interface_text_is_not_used_as_building_and_raw_display_is_preserved():
    displayed = (
        "東京都杉並区浜田山3-30-5 大きな地図を見る 周辺のお店を探す "
        "このお店は「杉並区浜田山4-14-7」から移転しています。"
    )
    source = _source(
        address_text_as_displayed=displayed,
        street_or_block="3-30-5",
        building="大きな地図を見る 周辺のお店を探す このお店は移転しています。",
    )
    result = _result(
        address_raw="東京都杉並区浜田山3-30-5",
        municipality_or_ward="杉並区",
        neighborhood="浜田山",
        street_or_block="3-30-5",
        source_evidence=[source],
    )
    agreement = compare_address_components(result)
    assert agreement.component_values["building"] == []
    assert source.address_text_as_displayed == displayed


def test_query_generation_prefers_japanese_name_and_ward_and_never_place_id():
    queries = generate_address_queries(_candidate())
    assert queries
    assert all("鮨さいとう" in query for query in queries[:3])
    assert all("Taito" in query for query in queries[:3])
    assert all("p1" not in query for query in queries)
    assert "Do not use the Google Place ID" in ADDRESS_RESEARCH_INSTRUCTIONS


def test_plan_only_selects_only_likely_missing_and_makes_no_api_calls(tmp_path):
    path = _db(tmp_path)
    before = path.stat().st_mtime_ns
    result = run_address_discovery(path, plan_only=True, limit=19)
    assert result["eligible_restaurant_count"] == 1
    assert [row["place_id"] for row in result["restaurants"]] == ["p1"]
    assert result["maximum_responses_requests"] == 1
    assert result["maximum_web_search_actions"] == 4
    assert result["compact_usage_controls"] == {
        "compact_research": True,
        "max_retained_sources": 4,
        "max_evidence_summary_chars": 160,
        "max_conflicting_candidates": 3,
        "max_output_tokens": 4000,
        "retry_truncated": False,
    }
    assert path.stat().st_mtime_ns == before


def test_read_only_osm_report_can_drive_initial_backfill_before_reason_migration(tmp_path):
    path = _db(tmp_path)
    with connect(path) as connection:
        connection.execute(
            "UPDATE public_restaurants SET location_resolution_reason=NULL WHERE place_id='p1'"
        )
        connection.commit()
    report_path = tmp_path / "osm-report.json"
    report_path.write_text(
        json.dumps(
            [
                {"place_id": "p1", "resolution_reason": "likely_not_represented_in_osm"},
                {
                    "place_id": "ambiguous",
                    "resolution_reason": "unresolved_ambiguous_exact_name_candidates",
                },
                {"place_id": "manual", "resolution_reason": "needs_manual_review"},
            ]
        ),
        encoding="utf-8",
    )
    result = run_address_discovery(
        path, plan_only=True, limit=19, resolution_report=report_path
    )
    assert result["eligible_restaurant_count"] == 1
    assert [row["place_id"] for row in result["restaurants"]] == ["p1"]


def test_paid_dry_run_uses_web_search_limit_tracks_usage_and_does_not_write(tmp_path):
    path = _db(tmp_path)
    before = deepcopy(_stored(path))
    client = FakeClient([_response(_result())])
    report = run_address_discovery(
        path,
        dry_run=True,
        place_id="p1",
        max_search_actions=2,
        client=client,
        model="configured-model",
    )
    call = client.responses.calls[0]
    assert call["tools"][0]["type"] == "web_search"
    assert call["max_tool_calls"] == 2
    assert call["max_output_tokens"] == 4000
    assert call["text"]["format"]["type"] == "json_schema"
    assert call["text"]["format"]["strict"] is True
    assert call["text"]["format"]["schema"]["additionalProperties"] is False
    assert "text_format" not in call
    assert report["usage_totals"]["response_request_count"] == 1
    assert report["usage_totals"]["web_search_action_count"] == 1
    assert report["usage_totals"]["cached_input_tokens"] == 25
    assert _stored(path) == before
    with connect(path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM address_research_runs").fetchone()[0] == 0


def test_complete_response_cut_off_mid_string_is_controlled_and_usage_is_preserved(tmp_path):
    path = _db(tmp_path)
    before = deepcopy(_stored(path))
    response = _response(
        None,
        raw_output='{"identity_status":"confirmed","research_summary":"cut off',
        response_status="completed",
    )
    report = run_address_discovery(
        path, dry_run=True, place_id="p1", max_search_actions=1,
        client=FakeClient([response]), model="configured-model",
    )
    failure = report["failures"][0]
    assert failure["failure_code"] == "malformed_structured_output"
    assert failure["parse_attempted"] is True
    assert failure["raw_output_character_count"] > 0
    assert "input_value" not in failure["failure_summary"]
    assert report["usage_totals"] == {
        "response_request_count": 1,
        "web_search_action_count": 1,
        "input_tokens": 100,
        "cached_input_tokens": 25,
        "output_tokens": 40,
        "reasoning_tokens": 10,
        "total_tokens": 140,
    }
    assert failure["web_search_action_audit"]["actual_action_count"] == 1
    assert failure["web_search_action_audit"]["counts_reconcile"] is True
    assert _stored(path) == before


def test_incomplete_max_output_response_is_not_parsed_or_retried_by_default(tmp_path):
    path = _db(tmp_path)
    response = _response(
        None,
        raw_output='{"identity_status":"confirmed","research_summary":"cut off',
        response_status="incomplete",
        incomplete_reason="max_output_tokens",
    )
    client = FakeClient([response, _response(_result())])
    report = run_address_discovery(
        path, dry_run=True, place_id="p1", max_search_actions=1,
        client=client, model="configured-model",
    )
    failure = report["failures"][0]
    assert failure["failure_code"] == "output_truncated"
    assert failure["response_status"] == "incomplete"
    assert failure["incomplete_reason"] == "max_output_tokens"
    assert failure["output_was_truncated"] is True
    assert failure["parse_attempted"] is False
    assert failure["usage"]["response_request_count"] == 1
    assert failure["usage"]["web_search_action_count"] == 1
    assert len(client.responses.calls) == 1


def test_parse_failure_persists_run_audit_but_no_evidence_or_location(tmp_path):
    path = _db(tmp_path)
    before = deepcopy(_stored(path))
    response = _response(
        None,
        raw_output='{"identity_status":"confirmed","research_summary":"cut off',
        response_status="incomplete",
        incomplete_reason="max_output_tokens",
    )
    report = run_address_discovery(
        path, place_id="p1", max_search_actions=1,
        client=FakeClient([response]), model="configured-model",
    )
    assert report["persisted"] == 0
    with connect(path) as connection:
        run = connection.execute("SELECT * FROM address_research_runs").fetchone()
        assert run["response_id"] == "resp_1"
        assert run["input_tokens"] == 100
        assert run["web_search_action_count"] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM address_search_attempts WHERE query_origin='actual_web_action'"
        ).fetchone()[0] == 1
        assert connection.execute("SELECT COUNT(*) FROM address_evidence").fetchone()[0] == 0
        assert connection.execute(
            "SELECT COUNT(*) FROM verified_restaurant_addresses"
        ).fetchone()[0] == 0
        assert connection.execute(
            "SELECT COUNT(*) FROM address_geocode_results"
        ).fetchone()[0] == 0
    after = _stored(path)
    assert after["map_display_eligible"] == before["map_display_eligible"]
    assert after["latitude"] == before["latitude"]
    assert after["longitude"] == before["longitude"]


def test_explicit_truncation_retry_is_bounded_and_aggregates_usage(tmp_path):
    path = _db(tmp_path)
    first = _response(
        None,
        raw_output='{"identity_status":"confirmed","research_summary":"cut off',
        response_status="incomplete",
        incomplete_reason="max_output_tokens",
        include_web_search=False,
        response_id="resp_truncated",
        usage_values={
            "input_tokens": 90, "cached_input_tokens": 20, "output_tokens": 4000,
            "reasoning_tokens": 100, "total_tokens": 4090,
        },
    )
    second = _response(
        _result(), response_id="resp_retry",
        usage_values={
            "input_tokens": 80, "cached_input_tokens": 10, "output_tokens": 500,
            "reasoning_tokens": 50, "total_tokens": 580,
        },
    )
    client = FakeClient([first, second])
    report = run_address_discovery(
        path, dry_run=True, place_id="p1", max_search_actions=1,
        retry_truncated=True, client=client, model="configured-model",
    )
    assert report["completed"] == 1
    assert len(client.responses.calls) == 2
    assert client.responses.calls[0]["max_output_tokens"] == 4000
    assert client.responses.calls[1]["max_output_tokens"] == 8000
    usage = report["restaurants"][0]["usage"]
    assert usage["response_ids"] == ["resp_truncated", "resp_retry"]
    assert usage["response_request_count"] == 2
    assert usage["retry_count"] == 1
    assert usage["attempts"][0]["response_status"] == "incomplete"
    assert usage["attempts"][0]["incomplete_reason"] == "max_output_tokens"
    assert usage["attempts"][1]["response_status"] == "completed"
    assert usage["input_tokens"] == 170
    assert usage["output_tokens"] == 4500
    assert usage["total_tokens"] == 4670
    assert report["usage_totals"]["response_request_count"] == 2
    assert report["usage_totals"]["web_search_action_count"] == 1


def test_retry_never_exceeds_cumulative_web_action_budget(tmp_path):
    path = _db(tmp_path)
    first = _response(
        None,
        raw_output='{"identity_status":"confirmed","research_summary":"cut off',
        response_status="incomplete",
        incomplete_reason="max_output_tokens",
    )
    client = FakeClient([first, _response(_result())])
    report = run_address_discovery(
        path, dry_run=True, place_id="p1", max_search_actions=1,
        retry_truncated=True, client=client, model="configured-model",
    )
    assert report["failed"] == 1
    assert len(client.responses.calls) == 1
    assert report["failures"][0]["web_search_action_audit"]["limit_reached"] is True


def test_strict_schema_unsupported_uses_audited_json_object_fallback(tmp_path):
    class UnsupportedStructuredOutput(RuntimeError):
        status_code = 400

    path = _db(tmp_path)
    client = FakeClient(
        [
            UnsupportedStructuredOutput("json_schema structured output is unsupported"),
            _response(_result()),
        ]
    )
    report = run_address_discovery(
        path, dry_run=True, place_id="p1", client=client, model="configured-model"
    )
    assert report["completed"] == 1
    assert len(client.responses.calls) == 2
    assert client.responses.calls[0]["text"]["format"]["type"] == "json_schema"
    assert client.responses.calls[1]["text"]["format"]["type"] == "json_object"
    assert report["restaurants"][0]["usage"]["structured_output_mode"] == "json_object_fallback"
    assert report["restaurants"][0]["usage"]["response_request_count"] == 2


def test_provider_action_budget_overrun_is_reported_rejected_and_stops_batch(tmp_path):
    path = _db(tmp_path)
    before = deepcopy(_stored(path))
    response = _response(_result())
    response.output.insert(
        1,
        SimpleNamespace(
            type="web_search_call",
            id="ws_2",
            status="completed",
            action=SimpleNamespace(type="open_page", queries=[], query=None, sources=[]),
        ),
    )
    report = run_address_discovery(
        path, dry_run=True, place_id="p1", max_search_actions=1,
        client=FakeClient([response]), model="configured-model",
    )
    row = report["restaurants"][0]
    assert row["acceptance"]["status"] == "failed"
    assert row["query_audit"]["configured_action_limit"] == 1
    assert row["query_audit"]["actual_action_count"] == 2
    assert row["query_audit"]["limit_reached"] is True
    assert row["query_audit"]["limit_exceeded"] is True
    assert row["query_audit"]["counts_reconcile"] is True
    assert report["usage_totals"]["web_search_action_count"] == 2
    assert report["failures"][0]["batch_stopped"] is True
    assert report["provider_completed"] == 1
    assert report["pipeline_accepted"] == 0
    assert report["pipeline_rejected"] == 1
    assert report["research_records_persisted"] == 0
    assert report["verified_addresses_persisted"] == 0
    assert report["unresolved_evidence_persisted"] == 0
    assert _stored(path) == before


def test_press_release_reclassification_and_tabelog_lead_policy(tmp_path):
    path = _db(tmp_path)
    result = _result(
        source_evidence=[
            _source(
                "restaurant_submission", source_url="https://prtimes.jp/main/html/example.html",
                restaurant_controlled=False,
            ),
            _source(
                "permitted_business_directory",
                source_url="https://tabelog.com/tokyo/example",
                restaurant_controlled=False,
            ),
        ]
    )
    report = run_address_discovery(
        path, dry_run=True, place_id="p1", client=FakeClient([_response(result)]),
        model="configured-model",
    )
    sources = report["restaurants"][0]["result"]["source_evidence"]
    assert sources[0]["source_type"] == "attributed_press_release"
    assert sources[0]["restaurant_controlled"] is False
    assert sources[1]["source_type"] == "lead_only_restricted_platform"
    assert sources[1]["supports_candidate_address"] is False
    tabelog_only = _result(
        source_evidence=[
            _source(
                "permitted_business_directory", source_url="https://tabelog.com/tokyo/example",
                restaurant_controlled=False,
            )
        ]
    )
    assert evaluate_address_result(_candidate(), tabelog_only).status == "provisional"


def test_compact_controls_limit_sources_summaries_and_conflicts(tmp_path):
    path = _db(tmp_path)
    verbose = "evidence " * 50
    result = _result(
        source_evidence=[
            _source(
                source_url=f"https://source{index}.example/access",
                identity_evidence_summary=verbose,
                address_evidence_summary=verbose,
            )
            for index in range(4)
        ],
        conflicting_address_candidates=[
            ConflictingAddressCandidate(
                address_raw=f"東京都台東区谷中1丁目2-{index}",
                source_urls=[f"https://conflict{index}.example"],
            )
            for index in range(4)
        ],
    )
    report = run_address_discovery(
        path, dry_run=True, place_id="p1", client=FakeClient([_response(result)]),
        model="configured-model", max_retained_sources=2,
        max_evidence_summary_chars=40, max_conflicting_candidates=1,
        max_output_tokens=900,
    )
    output = report["restaurants"][0]["result"]
    assert len(output["source_evidence"]) == 2
    assert len(output["source_evidence"][0]["identity_evidence_summary"]) == 40
    assert len(output["conflicting_address_candidates"]) == 1
    assert report["compact_usage_controls"]["max_output_tokens"] == 900


@pytest.mark.parametrize(
    "output",
    [
        SimpleNamespace(name="malformed"),
        RuntimeError("Responses API failed"),
        _response(_result(), search_status="failed"),
    ],
)
def test_malformed_response_api_error_and_web_search_error_are_safe(tmp_path, output):
    path = _db(tmp_path)
    before = deepcopy(_stored(path))
    client = FakeClient([output])
    report = run_address_discovery(
        path, dry_run=True, place_id="p1", client=client, model="configured-model"
    )
    assert report["failed"] == 1
    assert report["persisted"] == 0
    assert _stored(path) == before


def test_failed_attempt_is_cached_and_only_force_retries(tmp_path):
    path = _db(tmp_path)
    failed = run_address_discovery(
        path,
        place_id="p1",
        client=FakeClient([RuntimeError("temporary failure")]),
        model="configured-model",
    )
    assert failed["failed"] == 1
    planned = run_address_discovery(path, place_id="p1", plan_only=True)
    assert planned["eligible_restaurant_count"] == 0
    forced_client = FakeClient([_response(_result())])
    forced = run_address_discovery(
        path,
        place_id="p1",
        force=True,
        client=forced_client,
        model="configured-model",
    )
    assert forced["completed"] == 1
    assert len(forced_client.responses.calls) == 1


def test_persisted_usage_queries_and_address_do_not_touch_editorial_fields(tmp_path):
    path = _db(tmp_path)
    before = deepcopy(_stored(path))
    evidence_id = _persist(path, _result())
    after = _stored(path)
    assert evidence_id > 0
    for field in (
        "is_published",
        "fiyu_score",
        "fiyu_confidence",
        "why_fiyu",
        "food_tags_json",
        "signature_dishes_json",
        "evidence_json",
        "map_display_eligible",
        "latitude",
        "longitude",
    ):
        assert after[field] == before[field]
    assert after["address_resolution_status"] == "address_verified"
    with connect(path) as connection:
        run = connection.execute("SELECT * FROM address_research_runs").fetchone()
        assert run["response_id"] == "resp_1"
        assert run["web_search_action_count"] == 1
        assert run["total_tokens"] == 140
        origins = {
            row[0] for row in connection.execute(
                "SELECT query_origin FROM address_search_attempts"
            )
        }
        assert origins == {"fiyu_generated", "model_requested", "actual_web_action"}
    status = address_resolution_status(path)
    assert status["address_verified"] == 1
    assert status["usage_totals"]["response_request_count"] == 1
    assert status["source_type_distribution"]["official_restaurant_website"] == 1


def test_deterministic_recalculation_dry_run_makes_no_calls_or_writes(tmp_path, monkeypatch):
    path = _db(tmp_path)
    evidence_id = _persist(path, _result())
    with connect(path) as connection:
        connection.execute("DELETE FROM verified_restaurant_addresses WHERE public_restaurant_id='p1'")
        connection.execute(
            "UPDATE address_evidence SET acceptance_status='conflicting' WHERE id=?",
            (evidence_id,),
        )
        connection.commit()
    before = path.read_bytes()

    def forbidden(*args, **kwargs):
        raise AssertionError("network/provider code must not be called")

    from fiyu import address_research

    monkeypatch.setattr(address_research, "OpenAI", forbidden)
    report = recalculate_address_decisions(path, place_id="p1", dry_run=True)
    assert report["pipeline_accepted"] == 1
    assert report["responses_api_calls"] == 0
    assert report["web_search_calls"] == 0
    assert report["geocoder_calls"] == 0
    assert report["decision_audit_records_persisted"] == 0
    assert report["verified_addresses_persisted"] == 0
    assert path.read_bytes() == before


def test_recalculation_appends_decision_history_and_preserves_usage_and_evidence(tmp_path):
    path = _db(tmp_path)
    raw_display = "東京都台東区谷中１丁目２－３"
    result = _result(
        source_evidence=[
            _source(
                address_text_as_displayed=raw_display,
                street_or_block="１丁目２－３",
            )
        ]
    )
    evidence_id = _persist(path, result)
    with connect(path) as connection:
        connection.execute("DELETE FROM verified_restaurant_addresses WHERE public_restaurant_id='p1'")
        connection.execute(
            """
            UPDATE address_evidence SET acceptance_status='conflicting',
                acceptance_reasons_json='[\"material_component_conflict:street_or_block\"]'
            WHERE id=?
            """,
            (evidence_id,),
        )
        connection.commit()
        usage_before = dict(
            connection.execute(
                "SELECT * FROM address_research_runs WHERE id=(SELECT research_run_id FROM address_evidence WHERE id=?)",
                (evidence_id,),
            ).fetchone()
        )
        evidence_before = dict(
            connection.execute("SELECT * FROM address_evidence WHERE id=?", (evidence_id,)).fetchone()
        )

    first = recalculate_address_decisions(path, place_id="p1")
    second = recalculate_address_decisions(path, place_id="p1")
    assert first["pipeline_accepted"] == 1
    assert first["decision_audit_records_persisted"] == 1
    assert first["verified_addresses_persisted"] == 1
    assert second["decision_audit_records_persisted"] == 1
    assert second["verified_addresses_persisted"] == 0
    assert first["decisions"][0]["component_agreement"]["agreed_core_address"] == (
        "東京都台東区谷中1丁目2-3"
    )
    with connect(path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM address_decision_audits").fetchone()[0] == 2
        usage_after = dict(
            connection.execute("SELECT * FROM address_research_runs WHERE id=?", (usage_before["id"],)).fetchone()
        )
        evidence_after = dict(
            connection.execute("SELECT * FROM address_evidence WHERE id=?", (evidence_id,)).fetchone()
        )
        verified = connection.execute(
            "SELECT * FROM verified_restaurant_addresses WHERE public_restaurant_id='p1'"
        ).fetchone()
    assert usage_after == usage_before
    assert evidence_after == evidence_before
    assert json.loads(evidence_after["source_evidence_json"])[0]["address_text_as_displayed"] == raw_display
    assert verified["verified_core_address"] == "東京都台東区谷中1丁目2-3"


def test_accepted_combined_evidence_skips_fallback_and_missing_evidence_routes_to_it(tmp_path):
    path = _db(tmp_path)
    response = _response(_result())
    accepted_call = combined_address_call(
        _candidate(), result=_result(), response=response, model="configured-model"
    )
    run_id = start_address_run(
        path,
        place_id="p1",
        model="configured-model",
        forced=False,
        combined_research=True,
    )
    persist_address_call(
        path, place_id="p1", run_id=run_id, call=accepted_call
    )
    assert run_address_discovery(path, plan_only=True)["eligible_restaurant_count"] == 0

    second = _db(tmp_path / "second")
    missing_call = combined_address_call(
        _candidate(), result=None, response=_response(None), model="configured-model"
    )
    run_id = start_address_run(
        second,
        place_id="p1",
        model="configured-model",
        forced=False,
        combined_research=True,
    )
    persist_address_call(second, place_id="p1", run_id=run_id, call=missing_call)
    assert run_address_discovery(second, plan_only=True)["eligible_restaurant_count"] == 1


def test_ambiguous_combined_address_evidence_routes_to_standalone_fallback(tmp_path):
    path = _db(tmp_path)
    ambiguous = _result(identity_status="ambiguous", identity_confidence=0.6)
    call = combined_address_call(
        _candidate(), result=ambiguous, response=_response(ambiguous), model="configured-model"
    )
    assert call.acceptance.status == "conflicting"
    run_id = start_address_run(
        path,
        place_id="p1",
        model="configured-model",
        forced=False,
        combined_research=True,
    )
    persist_address_call(path, place_id="p1", run_id=run_id, call=call)
    assert run_address_discovery(path, plan_only=True)["eligible_restaurant_count"] == 1


def _research_payload(address=None):
    return {
        "matched_restaurant": True,
        "identity_confidence": 0.95,
        "name_ja": "鮨さいとう",
        "name_en": "Sushi Saito",
        "primary_category": "sushi",
        "food_tags": ["寿司"],
        "signature_dishes": ["おまかせ"],
        "official_language": "ja",
        "japanese_source_count": 3,
        "english_tourist_source_count": 1,
        "japanese_review_share": None,
        "tourist_coverage": "low",
        "reservation_platform_count": 1,
        "official_website_found": True,
        "social_profile_count": 1,
        "likely_chain": False,
        "known_location_count": 1,
        "specialist_restaurant": True,
        "independent_positive_source_count": 2,
        "total_evidence_sources": 4,
        "conflicting_evidence": False,
        "why_fiyu": "A focused sushi counter supported by independent local evidence.",
        "evidence_urls": ["https://editorial.example/saito"],
        **({"address_evidence": address} if address is not None else {}),
    }


def test_existing_research_without_address_is_backward_compatible_and_score_unchanged():
    without = RestaurantResearch.model_validate(_research_payload())
    with_address = RestaurantResearch.model_validate(
        _research_payload(_result().model_dump(mode="json"))
    )
    assert without.address_evidence is None
    internal = InternalSignals(80, 70, 60)
    assert calculate_fiyu_score(without.to_evidence(), internal) == calculate_fiyu_score(
        with_address.to_evidence(), internal
    )


def test_combined_future_research_uses_one_response_and_no_second_address_call(
    tmp_path, monkeypatch
):
    path = tmp_path / "combined.db"
    with connect(path) as connection:
        connection.executescript(SCHEMA)
        connection.commit()
    ensure_public_schema(path)
    with connect(path) as connection:
        connection.execute(
            """
            INSERT INTO restaurants (
                place_id, title, rating, review_count, candidate_eligible,
                internal_fiyu_score, confidence_score, quality_score,
                underexposure_score, digital_footprint_score, source_areas_json,
                score_reasons_json, source_files_json
            ) VALUES ('future', '鮨さいとう', 4.5, 20, 1, 80, 80, 70, 70, 70, '[]', '[]', '[]')
            """
        )
        connection.execute(
            """
            INSERT INTO public_restaurants (
                place_id, source_restaurant_id, discovery_area, discovery_area_type,
                discovery_areas_json, research_status, is_published, created_at, updated_at
            ) VALUES ('future', 1, 'Taito', 'ward', ?, 'pending', 0, 'created', 'before')
            """,
            (_candidate()["discovery_areas_json"],),
        )
        connection.commit()
    parsed = RestaurantResearch.model_validate(
        _research_payload(_result().model_dump(mode="json"))
    )
    client = FakeClient([_response(parsed)])
    import fiyu.research_worker as worker

    monkeypatch.setenv("OPENAI_API_KEY", "test-key-not-real")
    monkeypatch.setattr(worker, "load_dotenv", lambda: None)
    monkeypatch.setattr(worker, "OpenAI", lambda **kwargs: client)
    result = run_research_batch(path, limit=1, model="configured-model")
    assert result["completed"] == 1
    assert result["address_accepted"] == 1
    assert len(client.responses.calls) == 1
    with connect(path) as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM address_research_runs WHERE combined_research=1"
        ).fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM verified_restaurant_addresses"
        ).fetchone()[0] == 1


def test_review_export_dry_run_approval_and_rejection(tmp_path):
    path = _db(tmp_path)
    weak = _result(
        matched_name="別の鮨",
        source_evidence=[
            _source(
                "search_result_snippet",
                source_url="https://lead.example/saito",
                restaurant_controlled=False,
            )
        ]
    )
    evidence_id = _persist(path, weak)
    review = tmp_path / "address-review.csv"
    assert export_address_review(path, review) == 1
    with review.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    assert int(rows[0]["candidate_evidence_id"]) == evidence_id
    rows[0].update(
        reviewer_decision="approve",
        reviewer_notes="Checked the restaurant-controlled source independently.",
        reviewed_by="reviewer-1",
        reviewed_at="2026-07-27",
    )
    with review.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    path = _sqlite_backup_without_sidecars(path, tmp_path / "dry-run-clean.sqlite")
    before_dry_run = _sqlite_artifact_state(path)
    assert before_dry_run["-wal"] is None
    assert before_dry_run["-shm"] is None
    dry = import_address_review(path, review, dry_run=True)
    assert dry["updated"] == 0
    assert _sqlite_artifact_state(path) == before_dry_run
    with connect(path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM verified_restaurant_addresses").fetchone()[0] == 0
    applied = import_address_review(path, review)
    assert applied["updated"] == 1
    with connect(path) as connection:
        verified = connection.execute(
            "SELECT status, address_confidence_tier, decision_fingerprint "
            "FROM verified_restaurant_addresses"
        ).fetchone()
        assert verified["status"] == "address_verified"
        assert verified["address_confidence_tier"] == "manual"
        assert verified["decision_fingerprint"]

    rejected_db = _db(tmp_path / "rejected")
    _persist(rejected_db, weak)
    rejected_csv = tmp_path / "rejected.csv"
    export_address_review(rejected_db, rejected_csv)
    with rejected_csv.open(newline="", encoding="utf-8-sig") as handle:
        rejected_rows = list(csv.DictReader(handle))
    rejected_rows[0].update(
        reviewer_decision="reject", reviewed_by="reviewer", reviewed_at="2026-07-27"
    )
    with rejected_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=rejected_rows[0].keys())
        writer.writeheader()
        writer.writerows(rejected_rows)
    import_address_review(rejected_db, rejected_csv)
    with connect(rejected_db) as connection:
        assert connection.execute("SELECT COUNT(*) FROM verified_restaurant_addresses").fetchone()[0] == 0


def test_review_export_uses_latest_effective_real_batch_decisions(tmp_path):
    path = _db(tmp_path)
    restaurants = {
        "p1": ("あたらよ 秋葉原店", "Chiyoda"),
        "ambiguous": ("浜田山叙々苑", "Suginami"),
        "manual": ("江戸酒場 海", "Shibuya"),
    }
    with connect(path) as connection:
        for place_id, (name, area) in restaurants.items():
            connection.execute(
                """
                UPDATE public_restaurants SET name_ja=?, name_en=NULL,
                    discovery_area=?, discovery_areas_json=?
                WHERE place_id=?
                """,
                (
                    name,
                    area,
                    json.dumps([{"area": area, "area_type": "ward"}]),
                    place_id,
                ),
            )
        connection.commit()

    atarayo_candidate = _candidate(
        "p1", name_ja="あたらよ 秋葉原店", name_en=None, discovery_area="Chiyoda",
        discovery_areas_json=json.dumps([{"area": "Chiyoda", "area_type": "ward"}]),
    )
    atarayo_address = "東京都千代田区神田佐久間町1-13"
    atarayo = AddressResearchResult(
        identity_status="confirmed", identity_confidence=0.98,
        matched_name="あたらよ 秋葉原店", branch_name="秋葉原店",
        address_raw=atarayo_address, prefecture="東京都", municipality_or_ward="千代田区",
        neighborhood="神田佐久間町", street_or_block="1-13",
        source_evidence=[
            _source(
                "permitted_booking_platform", source_url="https://booking.example/atarayo",
                address_text_as_displayed=atarayo_address, municipality_or_ward="千代田区",
                neighborhood="神田佐久間町", street_or_block="1-13",
                restaurant_controlled=False,
            ),
            _source(
                "established_local_editorial_source",
                source_url="https://editorial.example/atarayo",
                address_text_as_displayed="東京都千代田区神田佐久間町１丁目１３",
                municipality_or_ward="千代田区", neighborhood="神田佐久間町",
                street_or_block="１丁目１３", restaurant_controlled=False,
            ),
        ],
    )
    atarayo_evidence = _persist_for_candidate(path, atarayo_candidate, atarayo)

    hamada_candidate = _candidate(
        "ambiguous", name_ja="浜田山叙々苑", name_en=None, discovery_area="Suginami",
        discovery_areas_json=json.dumps([{"area": "Suginami", "area_type": "ward"}]),
    )
    hamada_current = "東京都杉並区浜田山3-30-5"
    hamada_old = "東京都杉並区浜田山4-14-7"
    hamada = AddressResearchResult(
        identity_status="probable", identity_confidence=0.93, matched_name="浜田山叙々苑",
        address_raw=hamada_current, prefecture="東京都", municipality_or_ward="杉並区",
        neighborhood="浜田山", street_or_block="3-30-5",
        source_evidence=[
            _source(
                "permitted_booking_platform", source_url="https://booking.example/hamada",
                address_text_as_displayed=hamada_current, municipality_or_ward="杉並区",
                neighborhood="浜田山", street_or_block="3-30-5", restaurant_controlled=False,
            ),
            _source(
                "permitted_business_directory", source_url="https://directory.example/hamada-old",
                address_text_as_displayed=hamada_old, municipality_or_ward="杉並区",
                neighborhood="浜田山", street_or_block="4-14-7",
                building="大きな地図を見る 周辺のお店を探す このお店は移転しています。",
                address_evidence_summary="旧住所として表示。", restaurant_controlled=False,
            ),
        ],
        conflicting_address_candidates=[
            ConflictingAddressCandidate(
                address_raw=hamada_old, municipality_or_ward="杉並区",
                neighborhood="浜田山", street_or_block="4-14-7", summary="移転前の住所。",
            )
        ],
    )
    hamada_evidence = _persist_for_candidate(path, hamada_candidate, hamada)

    edo_candidate = _candidate(
        "manual", name_ja="江戸酒場 海", name_en=None, discovery_area="Shibuya",
        discovery_areas_json=json.dumps([{"area": "Shibuya", "area_type": "ward"}]),
    )
    edo_address = "東京都渋谷区神宮前2-23-4"
    edo = AddressResearchResult(
        identity_status="probable", identity_confidence=0.9, matched_name="江戸酒場 海",
        address_raw=edo_address, prefecture="東京都", municipality_or_ward="渋谷区",
        neighborhood="神宮前", street_or_block="2-23-4",
        source_evidence=[
            _source(
                "permitted_business_directory", source_url="https://directory.example/edo",
                address_text_as_displayed=edo_address, municipality_or_ward="渋谷区",
                neighborhood="神宮前", street_or_block="2-23-4", restaurant_controlled=False,
            )
        ],
    )
    edo_evidence = _persist_for_candidate(path, edo_candidate, edo)

    stale_conflict = json.dumps(
        {
            "agreed_core_address": None,
            "core_address_verified": False,
            "full_address_verified": False,
            "material_conflicting_components": ["street_or_block"],
            "non_material_conflicting_components": ["building"],
            "component_values": {
                "street_or_block": ["3-30-5", "4-14-7"],
                "building": ["大きな地図を見る 周辺のお店を探す"],
            },
            "proposed_location_precision": "unknown",
            "map_location_approximate": False,
        },
        ensure_ascii=False,
    )
    with connect(path) as connection:
        connection.execute("DELETE FROM verified_restaurant_addresses")
        connection.execute(
            "UPDATE address_evidence SET acceptance_status='conflicting', component_agreement_json=? WHERE id=?",
            (stale_conflict, atarayo_evidence),
        )
        connection.execute(
            "UPDATE address_evidence SET acceptance_status='conflicting', component_agreement_json=? WHERE id=?",
            (stale_conflict, hamada_evidence),
        )
        connection.execute(
            "UPDATE address_evidence SET acceptance_status='failed' WHERE id=?",
            (edo_evidence,),
        )
        connection.commit()

    atarayo_recalculated = recalculate_address_decisions(path, place_id="p1")
    hamada_recalculated = recalculate_address_decisions(path, place_id="ambiguous")
    edo_recalculated = recalculate_address_decisions(path, place_id="manual")
    output = tmp_path / "effective-address-review.csv"
    assert export_address_review(path, output) == 0
    assert atarayo_recalculated["decisions"][0]["proposed_acceptance"]["status"] == "provisional"
    assert hamada_recalculated["decisions"][0]["proposed_acceptance"]["status"] == "provisional"
    assert edo_recalculated["decisions"][0]["proposed_acceptance"]["status"] == "provisional"

    hamada_agreement = hamada_recalculated["decisions"][0]["component_agreement"]
    assert hamada_agreement["agreed_core_address"] == hamada_current
    assert hamada_agreement["material_conflicting_components"] == []
    assert hamada_agreement["street_or_block_agreement"] == "agrees"
    assert hamada_agreement["excluded_temporal_evidence"]
    assert all(
        item["address_text"] == hamada_old
        for item in hamada_agreement["excluded_temporal_evidence"]
    )
    hamada_reasons = hamada_recalculated["decisions"][0]["proposed_acceptance"]["reasons"]
    assert "material_component_conflict" not in " ".join(hamada_reasons)
    assert "identity_probable_but_consistent" in hamada_reasons

    edo_agreement = edo_recalculated["decisions"][0]["component_agreement"]
    assert edo_agreement["agreed_core_address"] == edo_address
    assert edo_agreement["material_conflicting_components"] == []


def test_review_rejects_duplicate_decisions_stale_evidence_and_invalid_date(tmp_path):
    path = _db(tmp_path)
    weak = _result(
        matched_name="別の鮨",
        source_evidence=[
            _source("search_result_snippet", restaurant_controlled=False)
        ]
    )
    _persist(path, weak)
    review = tmp_path / "invalid-review.csv"
    export_address_review(path, review)
    with review.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    rows[0].update(
        reviewer_decision="approve",
        reviewed_by="reviewer",
        reviewed_at="07/27/2026",
        evidence_fingerprint="stale",
    )
    rows.append(dict(rows[0]))
    with review.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    path = _sqlite_backup_without_sidecars(path, tmp_path / "duplicate-dry-run.sqlite")
    before_dry_run = _sqlite_artifact_state(path)
    result = import_address_review(path, review, dry_run=True)
    assert result["updated"] == 0
    assert result["validation_failures"] == 2
    assert any("duplicate restaurant decision" in error for error in result["reports"][1]["errors"])
    assert _sqlite_artifact_state(path) == before_dry_run

    persisted = import_address_review(path, review)
    assert persisted["updated"] == 0
    assert persisted["validation_failures"] == 2


def test_address_review_dry_run_rejects_old_schema_without_migrating(tmp_path):
    path = tmp_path / "old-schema.sqlite"
    connection = sqlite3.connect(path)
    try:
        connection.execute("CREATE TABLE public_restaurants (place_id TEXT PRIMARY KEY)")
        connection.commit()
    finally:
        connection.close()
    review = tmp_path / "empty-review.csv"
    review.write_text("public_restaurant_id,reviewer_decision\n", encoding="utf-8")
    before = _sqlite_artifact_state(path)

    with pytest.raises(RuntimeError, match="requires the current database schema") as error:
        import_address_review(path, review, dry_run=True)

    assert "public_cli --db PATH init" in str(error.value)
    assert _sqlite_artifact_state(path) == before


def test_review_export_ignores_superseded_evidence_and_manual_decision_takes_precedence(tmp_path):
    path = _db(tmp_path)
    weak = _result(
        matched_name="別の鮨",
        source_evidence=[
            _source("search_result_snippet", restaurant_controlled=False)
        ]
    )
    old_evidence = _persist(path, weak)
    recalculate_address_decisions(path, place_id="p1")
    new_evidence = _persist(path, weak)
    recalculate_address_decisions(path, place_id="p1")
    with connect(path) as connection:
        fingerprint = connection.execute(
            "SELECT evidence_fingerprint FROM address_evidence WHERE id=?", (new_evidence,)
        ).fetchone()[0]
        connection.execute(
            """
            INSERT INTO address_review_decisions (
                public_restaurant_id, address_evidence_id, reviewer_decision,
                reviewer_notes, reviewed_by, reviewed_at, evidence_fingerprint, created_at
            ) VALUES ('p1', ?, 'approve_core_location', 'Manual decision wins.',
                      'reviewer', '2026-07-28', ?, '2026-07-28T12:00:00+00:00')
            """,
            (new_evidence, fingerprint),
        )
        connection.execute(
            """
            INSERT INTO address_decision_audits (
                public_restaurant_id, address_evidence_id, decision_version,
                acceptance_status, resolution_status, acceptance_reasons_json,
                component_agreement_json, temporal_evidence_json,
                original_evidence_fingerprint, created_at
            )
            SELECT public_restaurant_id, id, 'later-test-audit', 'needs_review',
                   'address_needs_review', '[\"later deterministic decision\"]',
                   component_agreement_json, '[]', evidence_fingerprint,
                   '2026-07-28T13:00:00+00:00'
            FROM address_evidence WHERE id=?
            """,
            (new_evidence,),
        )
        connection.commit()
    review = tmp_path / "precedence.csv"
    assert export_address_review(path, review) == 0
    with connect(path) as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM address_decision_audits WHERE address_evidence_id=?",
            (old_evidence,),
        ).fetchone()[0] == 1


def test_xlsx_export_formats_address_fields_as_text_and_import_rejects_excel_date(tmp_path):
    from openpyxl import load_workbook

    path = _db(tmp_path)
    weak = _result(
        matched_name="別の鮨",
        address_raw="東京都台東区谷中1-13",
        street_or_block="1-13",
        source_evidence=[
            _source(
                "search_result_snippet", restaurant_controlled=False,
                address_text_as_displayed="東京都台東区谷中1-13", street_or_block="1-13",
            )
        ],
    )
    _persist(path, weak)
    review = tmp_path / "address-review.xlsx"
    assert export_address_review(path, review) == 1
    workbook = load_workbook(review)
    sheet = workbook.active
    headers = {cell.value: cell.column for cell in sheet[1]}
    street_cell = sheet.cell(2, headers["street_or_block"])
    assert street_cell.value == "1-13"
    assert street_cell.data_type == "s"
    assert street_cell.number_format == "@"
    sheet.cell(2, headers["street_or_block"], "Jan-13")
    sheet.cell(2, headers["reviewer_decision"], "approve_core_location")
    sheet.cell(2, headers["reviewed_by"], "reviewer")
    sheet.cell(2, headers["reviewed_at"], "2026-07-28")
    workbook.save(review)

    path = _sqlite_backup_without_sidecars(path, tmp_path / "excel-date-dry-run.sqlite")
    before = _sqlite_artifact_state(path)
    assert before["-wal"] is None
    assert before["-shm"] is None
    report = import_address_review(path, review, dry_run=True)
    assert report["updated"] == 0
    assert report["validation_failures"] == 1
    assert any(
        "spreadsheet date" in error for error in report["reports"][0]["errors"]
    )
    assert _sqlite_artifact_state(path) == before


def test_import_rejects_review_when_effective_decision_changed_after_export(tmp_path):
    path = _db(tmp_path)
    weak = _result(
        matched_name="別の鮨",
        source_evidence=[
            _source("search_result_snippet", restaurant_controlled=False)
        ]
    )
    evidence_id = _persist(path, weak)
    review = tmp_path / "stale-decision.csv"
    export_address_review(path, review)
    with review.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    rows[0].update(
        reviewer_decision="unresolved", reviewed_by="reviewer", reviewed_at="2026-07-28"
    )
    with connect(path) as connection:
        fingerprint = connection.execute(
            "SELECT evidence_fingerprint FROM address_evidence WHERE id=?", (evidence_id,)
        ).fetchone()[0]
        agreement = connection.execute(
            "SELECT component_agreement_json FROM address_evidence WHERE id=?", (evidence_id,)
        ).fetchone()[0]
        connection.execute(
            """
            INSERT INTO address_decision_audits (
                public_restaurant_id, address_evidence_id, decision_version,
                acceptance_status, resolution_status, acceptance_reasons_json,
                component_agreement_json, temporal_evidence_json,
                original_evidence_fingerprint, created_at
            ) VALUES ('p1', ?, 'test', 'accepted', 'address_verified', '[]', ?, '[]', ?,
                      '2099-01-01T00:00:00+00:00')
            """,
            (evidence_id, agreement, fingerprint),
        )
        connection.commit()
    with review.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    path = _sqlite_backup_without_sidecars(path, tmp_path / "stale-dry-run.sqlite")
    before = _sqlite_artifact_state(path)
    report = import_address_review(path, review, dry_run=True)
    assert report["updated"] == 0
    assert any(
        "effective_decision_fingerprint does not match" in error
        for error in report["reports"][0]["errors"]
    )
    assert any(
        "no longer has current effective status" in error
        for error in report["reports"][0]["errors"]
    )
    assert _sqlite_artifact_state(path) == before


def test_review_can_approve_core_without_approving_disputed_details(tmp_path):
    path = _db(tmp_path)
    core = "東京都台東区谷中1丁目2-3"
    lead = _source(
        "search_result_snippet",
        source_url="https://lead.example/saito",
        restaurant_controlled=False,
        address_text_as_displayed=f"{core} Building A 1階",
        building="Building A",
        floor="1階",
    )
    result = _result(
        matched_name="別の鮨",
        address_raw=f"{core} Building A 1階",
        building="Building A",
        floor="1階",
        source_evidence=[lead],
        conflicting_address_candidates=[
            ConflictingAddressCandidate(
                address_raw=f"{core} Building B B1F",
                source_urls=["https://lead2.example/saito"],
                building="Building B",
                floor="B1F",
            )
        ],
    )
    _persist(path, result)
    review = tmp_path / "core-review.csv"
    export_address_review(path, review)
    with review.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["proposed_decision"] == "approve_core_location"
    assert "building" in rows[0]["non_material_conflicting_components"]
    rows[0].update(
        reviewer_decision="approve_core_location", reviewed_by="reviewer",
        reviewed_at="2026-07-27",
    )
    with review.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    applied = import_address_review(path, review)
    assert applied["updated"] == 1
    with connect(path) as connection:
        verified = connection.execute("SELECT * FROM verified_restaurant_addresses").fetchone()
        assert verified["address_raw"] == core
        assert verified["geocoding_address"] == core
        assert verified["building"] is None and verified["floor"] is None
        assert verified["core_address_verified"] == 1
        assert verified["full_address_verified"] == 0


def _seed_verified_address(path, place_id="p1"):
    with connect(path) as connection:
        connection.execute(
            """
            INSERT INTO verified_restaurant_addresses (
                public_restaurant_id, address_raw, prefecture, municipality_or_ward,
                neighborhood, street_or_block, verified_core_address, geocoding_address,
                core_address_verified, full_address_verified, verification_method,
                evidence_references_json, verified_by, verified_at, status,
                created_at, updated_at
            ) VALUES (?, '東京都台東区谷中1丁目2-3', '東京都', '台東区', '谷中', '1丁目2-3',
                      '東京都台東区谷中1丁目2-3', '東京都台東区谷中1丁目2-3', 1, 1,
                      'manual_address_review', '["https://source.example"]', 'reviewer',
                      '2026-07-27', 'address_verified', 'created', 'updated')
            """,
            (place_id,),
        )
        connection.execute(
            "UPDATE public_restaurants SET address_resolution_status='address_verified' WHERE place_id=?",
            (place_id,),
        )
        connection.commit()


def _seed_core_only_conflict(path, place_id="p1"):
    core = "東京都台東区谷中1丁目2-3"
    with connect(path) as connection:
        connection.execute(
            """
            INSERT INTO verified_restaurant_addresses (
                public_restaurant_id, address_raw, prefecture, municipality_or_ward,
                neighborhood, street_or_block, verified_core_address, geocoding_address,
                core_address_verified, full_address_verified, unresolved_address_detail,
                approved_location_precision, map_location_approximate, verification_method,
                evidence_references_json, verified_by, verified_at, status, created_at, updated_at
            ) VALUES (?, ?, '東京都', '台東区', '谷中', '1丁目2-3', ?, ?, 1, 0,
                      'building: Building A vs Building B; floor: 1階 vs B1F',
                      'parcel_or_street_number', 1, 'manual_address_review',
                      '["https://source.example"]', 'reviewer', '2026-07-27',
                      'address_verified', 'created', 'updated')
            """,
            (place_id, core, core, core),
        )
        connection.execute(
            """
            UPDATE public_restaurants SET address_resolution_status='address_verified',
                verified_core_address=?, core_address_verified=1, full_address_verified=0,
                unresolved_address_detail='building/floor conflict'
            WHERE place_id=?
            """,
            (core, place_id),
        )
        connection.commit()


class MockGeocoder:
    def __init__(self, result):
        self.result = result
        self.calls = []

    def geocode(self, address, *, place_id=None, input_fingerprint=None):
        self.calls.append(address)
        return replace(
            self.result, place_id=place_id, input_fingerprint=input_fingerprint
        )


def _geocode(**changes):
    value = {
        "raw_address": "東京都台東区谷中1丁目2-3",
        "normalized_address": "東京都台東区谷中1-2-3",
        "latitude": 35.727,
        "longitude": 139.77,
        "prefecture": "東京都",
        "municipality_or_ward": "台東区",
        "address_level_match": "block",
        "precision": "exact",
        "provider": "independent_offline_geocoder",
        "provider_version": "fixture-v1",
        "source_reference": "file:///reviewed/geocode.json",
    }
    value.update(changes)
    return AddressGeocodeResult(**value)


def test_only_verified_addresses_are_geocoded_and_dry_run_does_not_write(tmp_path):
    path = _db(tmp_path)
    geocoder = MockGeocoder(_geocode())
    assert geocode_verified_addresses(path, geocoder=geocoder, dry_run=True)["selected"] == 0
    assert not geocoder.calls
    _seed_verified_address(path)
    before = deepcopy(_stored(path))
    before_artifacts = _sqlite_artifact_state(path)
    result = geocode_verified_addresses(path, geocoder=geocoder, dry_run=True)
    assert result["location_verified"] == 1
    assert len(geocoder.calls) == 1
    assert _sqlite_artifact_state(path) == before_artifacts
    assert _stored(path) == before
    with connect(path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM address_geocode_results").fetchone()[0] == 0


def test_core_only_geocoding_excludes_disputed_building_and_floor(tmp_path):
    path = _db(tmp_path)
    _seed_core_only_conflict(path)
    geocoder = MockGeocoder(_geocode(address_level_match="address"))
    result = geocode_verified_addresses(path, geocoder=geocoder)
    assert geocoder.calls == ["東京都台東区谷中1丁目2-3"]
    assert "Building" not in geocoder.calls[0] and "B1F" not in geocoder.calls[0]
    assert result["location_verified"] == 1
    stored = _stored(path)
    assert stored["map_display_eligible"] == 1
    assert stored["map_location_precision"] == "parcel_or_street_number"
    assert stored["map_location_approximate"] == 1
    assert stored["full_address_verified"] == 0


@pytest.mark.parametrize(
    ("match_level", "expected_status", "precision", "approximate"),
    [
        ("address", "location_verified", "parcel_or_street_number", False),
        ("block", "location_verified", "block", True),
        ("neighborhood", "geocode_needs_review", "neighborhood", False),
    ],
)
def test_map_precision_rules(match_level, expected_status, precision, approximate):
    validation = validate_geocode(_geocode(address_level_match=match_level), verified_ward="台東区")
    assert validation.status == expected_status
    assert validation.location_precision == precision
    assert validation.map_location_approximate is approximate


@pytest.mark.parametrize(
    ("result", "reason"),
    [
        (_geocode(address_level_match="ward", precision="area_anchor"), "geocode_match_level_not_street_detail"),
        (_geocode(municipality_or_ward="杉並区"), "geocoded_ward_mismatch"),
        (_geocode(latitude=40.0, longitude=139.77), "coordinates_outside_tokyo_bounds"),
        (_geocode(latitude=139.77, longitude=35.727), "latitude_longitude_appear_swapped"),
    ],
)
def test_low_precision_ward_mismatch_bounds_and_swapped_require_review(result, reason):
    validation = validate_geocode(result, verified_ward="台東区")
    assert validation.status == "geocode_needs_review"
    assert reason in validation.reasons


def test_valid_geocode_alone_sets_map_eligibility_and_preserves_kin(tmp_path):
    path = _db(tmp_path)
    kin_before = deepcopy(_stored(path, "ChIJ2WzWhfWPGGARyYQS7SD2tIM"))
    _seed_verified_address(path)
    result = geocode_verified_addresses(path, geocoder=MockGeocoder(_geocode()))
    assert result["location_verified"] == 1
    restaurant = _stored(path)
    assert restaurant["map_display_eligible"] == 1
    assert restaurant["location_verification_status"] == "location_verified"
    assert restaurant["location_source"] == "independent_offline_geocoder"
    assert _stored(path, "ChIJ2WzWhfWPGGARyYQS7SD2tIM") == kin_before
    with connect(path) as connection:
        history = connection.execute(
            "SELECT location_source, map_display_eligible FROM location_history "
            "WHERE public_restaurant_id='p1'"
        ).fetchall()
    assert [(row["location_source"], row["map_display_eligible"]) for row in history] == [
        ("independent_offline_geocoder", 1)
    ]


def test_area_fallback_persists_as_map_eligible_provisional_location(tmp_path):
    path = _db(tmp_path)
    kin_before = deepcopy(_stored(path, "ChIJ2WzWhfWPGGARyYQS7SD2tIM"))
    _seed_verified_address(path)
    area_result = _geocode(
        address_level_match="chome",
        precision="approximate",
        neighborhood="谷中",
        match_status="matched_chome_area_approximate",
        map_location_approximate=True,
        map_anchor_type="chome",
        matched_components={"ward": "台東区", "neighborhood": "谷中", "chome": "1"},
        unmatched_components={"block": "2", "sub_number": "3"},
        suggested_verification_tier="provisional_medium",
        osm_type="relation",
        osm_id=1234,
        osm_version=8,
        osm_timestamp="2026-07-01T00:00:00Z",
        representative_point_method="polygon_centroid_inside",
        provenance="Map data © OpenStreetMap contributors",
        source_reference="https://www.openstreetmap.org/relation/1234",
    )
    report = geocode_verified_addresses(path, geocoder=MockGeocoder(area_result))
    assert report["location_provisional"] == 1
    stored = _stored(path)
    assert stored["map_display_eligible"] == 1
    assert stored["location_verification_status"] == "location_provisional"
    assert stored["location_status"] == "location_provisional"
    assert stored["map_location_approximate"] == 1
    assert stored["map_location_precision"] == "chome"
    assert stored["map_anchor_type"] == "chome"
    assert json.loads(stored["location_unmatched_components_json"])["block"] == "2"
    assert _stored(path, "ChIJ2WzWhfWPGGARyYQS7SD2tIM") == kin_before
    with connect(path) as connection:
        history = connection.execute(
            "SELECT * FROM location_history WHERE public_restaurant_id='p1'"
        ).fetchone()
    assert history["location_status"] == "location_provisional"
    assert history["map_anchor_type"] == "chome"


def test_json_loader_retains_all_area_fallback_statuses_for_dry_run(tmp_path):
    path = _db(tmp_path)
    with connect(path) as connection:
        connection.executemany(
            """
            INSERT INTO public_restaurants (
                place_id, research_status, is_published,
                location_verification_status, map_display_eligible,
                address_resolution_status, created_at, updated_at
            ) VALUES (?, 'complete', 1, 'unresolved', 0,
                      'address_not_researched', 'created', 'before')
            """,
            [("area-block",), ("area-neighborhood",)],
        )
        connection.commit()
    for place_id in ("p1", "area-block", "area-neighborhood"):
        _seed_verified_address(path, place_id)

    inputs_path = tmp_path / "inputs.json"
    assert export_geocoding_inputs(path, inputs_path) == 3
    inputs = {
        item["place_id"]: item
        for item in json.loads(inputs_path.read_text(encoding="utf-8"))
    }
    status_by_place = {
        "area-block": ("matched_block_area_approximate", "block"),
        "p1": ("matched_chome_area_approximate", "chome"),
        "area-neighborhood": (
            "matched_neighborhood_area_approximate", "neighborhood"
        ),
    }
    payload = []
    for offset, (place_id, (status, level)) in enumerate(
        status_by_place.items(), start=1
    ):
        item = inputs[place_id]
        payload.append({
            "normalized_address": "台東区谷中1-2-3",
            "latitude": 35.727 + offset / 10000,
            "longitude": 139.77 + offset / 10000,
            "precision": "approximate",
            "warnings": [f"osm_{level}_area_fallback_is_approximate"],
            "provenance": "Map data © OpenStreetMap contributors",
            "raw_address": item["accepted_core_address"],
            "prefecture": "東京都",
            "municipality_or_ward": "台東区",
            "provider": "local_osm_addresses",
            "provider_version": "osm-address-index-v3-area",
            "source_reference": f"https://www.openstreetmap.org/relation/{9000 + offset}",
            "place_id": place_id,
            "input_fingerprint": item["input_fingerprint"],
            "neighborhood": "谷中",
            "matched_components": {"ward": "台東区", "neighborhood": "谷中"},
            "unmatched_components": {"block": "2", "sub_number": "3"},
            "match_status": status,
            "map_location_approximate": True,
            "suggested_verification_tier": "provisional_medium",
            "osm_type": "relation",
            "osm_id": 9000 + offset,
            "osm_version": 1,
            "osm_timestamp": "2026-07-01T00:00:00Z",
            "representative_point_method": "polygon_centroid_inside",
            "map_anchor_type": level,
            "diagnostic_candidates": [],
            "match_level": level,
            "status": status,
        })
    payload.extend([
        {
            "place_id": "rejected",
            "raw_address": "東京都台東区谷中9-9",
            "status": "rejected_address_number_mismatch",
            "match_status": "rejected_address_number_mismatch",
        },
        {
            "place_id": "malformed",
            "raw_address": "東京都台東区谷中8-8",
            "status": "matched_chome_area_approximate",
            "match_status": "matched_chome_area_approximate",
            "match_level": "chome",
            "longitude": 139.77,
        },
    ])
    results_path = tmp_path / "area-results.json"
    results_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    geocoder = JsonFileAddressGeocoder(results_path)
    assert geocoder.loaded_count == 3
    for place_id, (status, _level) in status_by_place.items():
        loaded = geocoder.geocode(
            str(inputs[place_id]["accepted_core_address"]), place_id=place_id
        )
        assert loaded is not None
        assert loaded.match_status == status
    assert [item["reason"] for item in geocoder.excluded_records] == [
        "status_not_importable:rejected_address_number_mismatch",
        "malformed_geocoder_result:KeyError:'latitude'",
    ]

    report = geocode_verified_addresses(path, geocoder=geocoder, dry_run=True)
    assert {
        key: report[key]
        for key in (
            "selected", "location_verified", "location_provisional",
            "needs_review", "failed",
        )
    } == {
        "selected": 3,
        "location_verified": 0,
        "location_provisional": 3,
        "needs_review": 0,
        "failed": 0,
    }


@pytest.mark.parametrize(
    ("record", "expected_reason"),
    [
        (
            {"status": "rejected_address_number_mismatch"},
            "status_not_importable:rejected_address_number_mismatch",
        ),
        (
            {
                "status": "matched_chome_area_approximate",
                "match_status": "matched_chome_area_approximate",
                "match_level": "chome",
                "longitude": 139.77,
            },
            "malformed_geocoder_result:KeyError:'latitude'",
        ),
    ],
)
def test_json_loader_reports_explicit_reason_for_excluded_selected_record(
    tmp_path, record, expected_reason
):
    path = _db(tmp_path)
    _seed_verified_address(path)
    inputs_path = tmp_path / "inputs.json"
    export_geocoding_inputs(path, inputs_path)
    item = json.loads(inputs_path.read_text(encoding="utf-8"))[0]
    results_path = tmp_path / "excluded.json"
    results_path.write_text(
        json.dumps([{
            "place_id": item["place_id"],
            "raw_address": item["accepted_core_address"],
            "input_fingerprint": item["input_fingerprint"],
            **record,
        }], ensure_ascii=False),
        encoding="utf-8",
    )
    report = geocode_verified_addresses(
        path, geocoder=JsonFileAddressGeocoder(results_path), dry_run=True
    )
    assert report["failed"] == 1
    assert report["reports"][0]["reasons"] == [expected_reason]


def test_invalid_geocode_is_persisted_for_review_without_map_eligibility(tmp_path):
    path = _db(tmp_path)
    _seed_verified_address(path)
    before = deepcopy(_stored(path))
    result = geocode_verified_addresses(
        path,
        geocoder=MockGeocoder(_geocode(address_level_match="ward", precision="area_anchor")),
    )
    after = _stored(path)
    assert result["needs_review"] == 1
    assert after["map_display_eligible"] == 0
    for field in ("is_published", "fiyu_score", "why_fiyu", "food_tags_json", "signature_dishes_json"):
        assert after[field] == before[field]
    with connect(path) as connection:
        row = connection.execute("SELECT * FROM address_geocode_results").fetchone()
        assert row["validation_status"] == "geocode_needs_review"


def test_stale_geocoding_input_fingerprint_cannot_set_map_eligibility(tmp_path):
    path = _db(tmp_path)
    _seed_verified_address(path)

    class StaleResultGeocoder:
        def geocode(self, address, **kwargs):
            return replace(
                _geocode(raw_address=address),
                place_id=kwargs["place_id"],
                input_fingerprint="stale-fingerprint",
            )

    before = _sqlite_artifact_state(path)
    result = geocode_verified_addresses(path, geocoder=StaleResultGeocoder())
    assert result["needs_review"] == 1
    assert "stale_or_missing_geocoding_input_fingerprint" in result["reports"][0]["reasons"]
    assert _sqlite_artifact_state(path) == before
    assert _stored(path)["map_display_eligible"] == 0
    with connect(path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM address_geocode_results").fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM location_history").fetchone()[0] == 0


def test_geocode_persistence_rolls_back_all_writes_on_history_failure(tmp_path):
    path = _db(tmp_path)
    _seed_verified_address(path)
    before = deepcopy(_stored(path))
    with connect(path) as connection:
        connection.execute(
            """
            CREATE TRIGGER fail_test_location_history
            BEFORE INSERT ON location_history
            BEGIN
                SELECT RAISE(ABORT, 'synthetic history failure');
            END
            """
        )
        connection.commit()

    with pytest.raises(sqlite3.IntegrityError, match="synthetic history failure"):
        geocode_verified_addresses(path, geocoder=MockGeocoder(_geocode()))

    assert _stored(path) == before
    with connect(path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM address_geocode_results").fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM location_history").fetchone()[0] == 0
        assert connection.execute(
            "SELECT status FROM verified_restaurant_addresses WHERE public_restaurant_id='p1'"
        ).fetchone()[0] == "address_verified"


def test_provisional_address_exports_core_and_abr_parcel_result_becomes_map_pin(tmp_path):
    from fiyu.address_geocoder import JsonFileAddressGeocoder

    path = _db(tmp_path)
    result = _result(
        address_raw="東京都台東区谷中1丁目2-3 ビルA 1階",
        building="ビルA", floor="1階",
        source_evidence=[
            _source(
                "permitted_business_directory",
                source_url="https://directory.example.jp/saito",
                restaurant_controlled=False,
            )
        ],
    )
    _persist(path, result)
    inputs = tmp_path / "geocoding-inputs.json"
    assert export_geocoding_inputs(path, inputs) == 1
    item = json.loads(inputs.read_text(encoding="utf-8"))[0]
    assert item["accepted_core_address"] == "東京都台東区谷中1丁目2-3"
    assert "ビル" not in item["accepted_core_address"]
    assert item["address_status"] == "address_provisionally_accepted"
    assert item["confidence_tier"] == "provisional_high"
    assert item["input_fingerprint"]

    abr_dir = tmp_path / "abr-data"
    abr_dir.mkdir()

    def fake_abr_runner(*args, **kwargs):
        return SimpleNamespace(
            returncode=0,
            stderr="",
            stdout=json.dumps(
                [
                    {
                        "query": {"input": item["accepted_core_address"]},
                        "result": {
                            "output": "東京都台東区谷中一丁目2-3",
                            "match_level": "residential_detail",
                            "coordinate_level": "residential_detail",
                            "lat": 35.727,
                            "lon": 139.77,
                            "pref": "東京都",
                            "city": "台東区",
                            "ward": None,
                            "machiaza_id": "fixture-town",
                            "rsdt_id": "fixture-address",
                        },
                    }
                ],
                ensure_ascii=False,
            ),
        )

    geocoder = DigitalAgencyAbrGeocoder(data_dir=abr_dir, runner=fake_abr_runner)
    results_path = tmp_path / "abr-results.json"
    batch = geocode_address_file(inputs, results_path, geocoder=geocoder)
    assert batch["geocoded"] == 1 and batch["failed"] == 0
    exported_result = json.loads(results_path.read_text(encoding="utf-8"))[0]
    assert exported_result["provider"] == "digital_agency_address_base_registry"
    assert exported_result["input_fingerprint"] == item["input_fingerprint"]

    imported = geocode_verified_addresses(
        path, geocoder=JsonFileAddressGeocoder(results_path), dry_run=True
    )
    assert imported["location_provisional"] == 1
    assert _stored(path)["map_display_eligible"] == 0
    persisted = geocode_verified_addresses(path, geocoder=JsonFileAddressGeocoder(results_path))
    assert persisted["location_provisional"] == 1
    stored = _stored(path)
    assert stored["map_display_eligible"] == 1
    assert stored["location_verification_status"] == "location_provisional"
    assert stored["location_verification_tier"] == "provisional_high"
    assert stored["map_location_approximate"] == 0


def test_provisional_medium_parcel_result_is_still_marked_approximate(tmp_path):
    path = _db(tmp_path)
    _persist(
        path,
        _result(
            identity_status="probable",
            identity_confidence=0.78,
            source_evidence=[
                _source(
                    "permitted_business_directory",
                    source_url="https://directory.example.jp/saito",
                    restaurant_controlled=False,
                )
            ],
        ),
    )
    report = geocode_verified_addresses(
        path,
        geocoder=MockGeocoder(
            _geocode(address_level_match="parcel", precision="exact")
        ),
    )
    assert report["location_provisional"] == 1
    assert report["reports"][0]["map_location_approximate"] is True
    stored = _stored(path)
    assert stored["location_verification_tier"] == "provisional_medium"
    assert stored["location_precision"] == "approximate"
    assert stored["map_location_approximate"] == 1
    status = address_resolution_status(path)
    assert status["provisional_map_locations"] == 1
    assert status["approximate_map_locations"] == 1
    assert status["geocoding_pending"] == 0


def test_abr_adapter_uses_coordinate_level_for_pin_precision(tmp_path):
    data_dir = tmp_path / "abr-data"
    data_dir.mkdir()

    def runner(*args, **kwargs):
        return SimpleNamespace(
            returncode=0,
            stderr="",
            stdout=json.dumps(
                [{
                    "result": {
                        "output": "東京都台東区谷中一丁目2-3",
                        "match_level": "residential_detail",
                        "coordinate_level": "residential_block",
                        "lat": 35.727,
                        "lon": 139.77,
                        "pref": "東京都",
                        "city": "台東区",
                    }
                }],
                ensure_ascii=False,
            ),
        )

    result = DigitalAgencyAbrGeocoder(data_dir=data_dir, runner=runner).geocode(
        "東京都台東区谷中1丁目2-3"
    )
    assert result is not None
    assert result.address_level_match == "block"
    assert result.precision == "approximate"
    assert "abr_coordinate_level:residential_block" in result.warnings


def test_local_osm_address_result_import_preserves_kin_and_full_provenance(tmp_path):
    from fiyu.address_geocoder import JsonFileAddressGeocoder, LocalOSMAddressGeocoder
    from fiyu.osm_index import OSMFeature, build_osm_index

    path = _db(tmp_path)
    kin_before = deepcopy(_stored(path, "ChIJ2WzWhfWPGGARyYQS7SD2tIM"))
    _persist(
        path,
        _result(
            source_evidence=[
                _source(
                    "permitted_business_directory",
                    source_url="https://directory.example.jp/saito",
                    restaurant_controlled=False,
                )
            ]
        ),
    )
    index = tmp_path / "osm-addresses.sqlite"
    build_osm_index(
        tmp_path / "synthetic.osm.pbf",
        index,
        features=[
            OSMFeature(
                "node",
                1234,
                {
                    "addr:prefecture": "東京都",
                    "addr:city": "台東区",
                    "addr:neighbourhood": "谷中一丁目",
                    "addr:housenumber": "2-3",
                },
                ((35.727, 139.77),),
                osm_version=5,
                osm_timestamp="2026-07-01T00:00:00Z",
            )
        ],
    )
    inputs = tmp_path / "inputs.json"
    results = tmp_path / "results.json"
    assert export_geocoding_inputs(path, inputs) == 1
    geocode_address_file(
        inputs,
        results,
        geocoder=LocalOSMAddressGeocoder(index),
    )
    before = path.read_bytes()
    preview = geocode_verified_addresses(
        path,
        geocoder=JsonFileAddressGeocoder(results),
        dry_run=True,
    )
    assert preview["location_provisional"] == 1
    assert path.read_bytes() == before
    applied = geocode_verified_addresses(
        path,
        geocoder=JsonFileAddressGeocoder(results),
    )
    assert applied["location_provisional"] == 1
    stored = _stored(path)
    assert stored["map_display_eligible"] == 1
    assert stored["location_source"] == "local_osm_addresses"
    assert stored["location_verification_tier"] == "provisional_high"
    assert stored["map_location_approximate"] == 0
    assert stored["location_osm_type"] == "node"
    assert stored["location_osm_id"] == 1234
    assert stored["location_osm_version"] == 5
    assert _stored(path, "ChIJ2WzWhfWPGGARyYQS7SD2tIM") == kin_before
    with connect(path) as connection:
        geocode = connection.execute(
            "SELECT * FROM address_geocode_results WHERE public_restaurant_id='p1'"
        ).fetchone()
    assert geocode["match_status"] == "matched_exact"
    assert geocode["osm_id"] == 1234
    assert json.loads(geocode["matched_components_json"])["neighborhood"] == "谷中"


def test_local_osm_not_found_restaurant_remains_published_without_pin(tmp_path):
    from fiyu.address_geocoder import JsonFileAddressGeocoder, LocalOSMAddressGeocoder
    from fiyu.osm_index import OSMFeature, build_osm_index

    path = _db(tmp_path)
    _persist(
        path,
        _result(
            source_evidence=[
                _source(
                    "permitted_business_directory",
                    source_url="https://directory.example.jp/saito",
                    restaurant_controlled=False,
                )
            ]
        ),
    )
    index = tmp_path / "empty-addresses.sqlite"
    build_osm_index(
        tmp_path / "synthetic.osm.pbf",
        index,
        features=[OSMFeature("node", 1, {"name": "unrelated"}, ((35.7, 139.7),))],
    )
    inputs = tmp_path / "inputs.json"
    results = tmp_path / "results.json"
    export_geocoding_inputs(path, inputs)
    batch = geocode_address_file(
        inputs,
        results,
        geocoder=LocalOSMAddressGeocoder(index),
    )
    assert batch["status_counts"] == {"not_found": 1}
    imported = geocode_verified_addresses(
        path,
        geocoder=JsonFileAddressGeocoder(results),
    )
    assert imported["failed"] == 1
    stored = _stored(path)
    assert stored["is_published"] == 1
    assert stored["map_display_eligible"] == 0
    assert stored["latitude"] is None and stored["longitude"] is None


def test_block_and_narrow_interpolation_are_approximate_but_broad_interpolation_rejected():
    block = validate_geocode(
        _geocode(address_level_match="block", precision="approximate"),
        verified_ward="台東区",
    )
    assert block.status == "location_verified" and block.map_location_approximate
    narrow = validate_geocode(
        _geocode(
            address_level_match="interpolation", precision="approximate",
            interpolation_span_meters=80,
        ),
        verified_ward="台東区",
    )
    assert narrow.status == "location_verified" and narrow.map_location_approximate
    broad = validate_geocode(
        _geocode(
            address_level_match="interpolation", precision="approximate",
            interpolation_span_meters=300,
        ),
        verified_ward="台東区",
    )
    assert broad.status == "geocode_needs_review"
    assert "geocode_interpolation_not_sufficiently_narrow" in broad.reasons


def test_geocoder_batch_isolates_one_failed_restaurant(tmp_path):
    inputs = [
        {"place_id": "one", "accepted_core_address": "東京都台東区谷中1-1",
         "input_fingerprint": "fingerprint-one"},
        {"place_id": "two", "accepted_core_address": "東京都台東区谷中1-2",
         "input_fingerprint": "fingerprint-two"},
    ]
    input_path = tmp_path / "inputs.json"
    output_path = tmp_path / "results.json"
    input_path.write_text(json.dumps(inputs, ensure_ascii=False), encoding="utf-8")

    class IsolatedFailureGeocoder:
        def geocode(self, address, **kwargs):
            if address.endswith("1-2"):
                raise RuntimeError("fixture failure")
            return _geocode(raw_address=address)

    report = geocode_address_file(
        input_path, output_path, geocoder=IsolatedFailureGeocoder()
    )
    assert report["selected"] == 2
    assert report["geocoded"] == 1
    assert report["failed"] == 1
    assert json.loads(output_path.read_text(encoding="utf-8"))[0]["place_id"] == "one"


def test_location_correction_preserves_history_and_manual_location_requires_override(tmp_path):
    path = _db(tmp_path)
    kin_before = deepcopy(_stored(path, "ChIJ2WzWhfWPGGARyYQS7SD2tIM"))
    kin_blocked = replace_location(
        path, place_id="ChIJ2WzWhfWPGGARyYQS7SD2tIM", latitude=35.72,
        longitude=139.79, source_reference="https://example.jp/correction",
        reason="Should require override", reviewed_by="Ethan", reviewed_at="2026-07-28",
        dry_run=True,
    )
    assert not kin_blocked["valid"]
    assert _stored(path, "ChIJ2WzWhfWPGGARyYQS7SD2tIM") == kin_before
    _seed_verified_address(path)
    geocode_verified_addresses(path, geocoder=MockGeocoder(_geocode()))
    before = deepcopy(_stored(path))
    dry = replace_location(
        path, place_id="p1", latitude=35.728, longitude=139.771,
        source_reference="https://example.jp/correction", reason="Corrected location",
        reviewed_by="Ethan", reviewed_at="2026-07-28", dry_run=True,
    )
    assert dry["valid"] and not dry["updated"]
    assert _stored(path) == before
    applied = replace_location(
        path, place_id="p1", latitude=35.728, longitude=139.771,
        source_reference="https://example.jp/correction", reason="Corrected location",
        reviewed_by="Ethan", reviewed_at="2026-07-28",
    )
    assert applied["updated"]
    with connect(path) as connection:
        history = connection.execute(
            "SELECT location_status, latitude FROM location_history WHERE public_restaurant_id='p1' ORDER BY id"
        ).fetchall()
    assert any(row["location_status"] == "location_superseded" for row in history)
    assert history[-1]["location_status"] == "location_active"
    assert _stored(path)["latitude"] == 35.728

    blocked = replace_location(
        path, place_id="p1", latitude=35.729, longitude=139.772,
        source_reference="https://example.jp/second", reason="Second correction",
        reviewed_by="Ethan", reviewed_at="2026-07-28", dry_run=True,
    )
    assert not blocked["valid"]
    assert "--allow-manual-override" in blocked["errors"][0]
    removed = replace_location(
        path, place_id="p1", latitude=None, longitude=None,
        source_reference="https://example.jp/remove", reason="Remove provisional pin",
        reviewed_by="Ethan", reviewed_at="2026-07-28", remove=True,
        allow_manual_override=True,
    )
    assert removed["updated"]
    stored = _stored(path)
    assert stored["map_display_eligible"] == 0
    assert stored["latitude"] is None and stored["longitude"] is None
    status = address_resolution_status(path)
    assert status["corrected_or_superseded_locations"] == 1
    assert status["corrections_count"] == 1
    assert status["locations_removed_from_map_display"] == 1
