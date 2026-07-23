from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Iterator, Mapping

from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".json", ".jsonl", ".ndjson", ".xlsx"}


def iter_input_files(paths: list[str | Path]) -> list[Path]:
    found: list[Path] = []
    for value in paths:
        path = Path(value)
        if path.is_dir():
            found.extend(
                child
                for child in sorted(path.rglob("*"))
                if child.is_file() and child.suffix.casefold() in SUPPORTED_EXTENSIONS
            )
        elif path.is_file() and path.suffix.casefold() in SUPPORTED_EXTENSIONS:
            found.append(path)
        else:
            raise FileNotFoundError(f"Unsupported or missing input: {path}")
    unique = sorted({item.resolve() for item in found})
    if not unique:
        raise FileNotFoundError("No supported dataset files found")
    return unique


def _iter_delimited(path: Path, delimiter: str) -> Iterator[Mapping[str, object]]:
    csv.field_size_limit(min(2_147_483_647, 1024 * 1024 * 1024))
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if not reader.fieldnames:
            return
        for row in reader:
            yield row


def _iter_json(path: Path) -> Iterator[Mapping[str, object]]:
    if path.suffix.casefold() in {".jsonl", ".ndjson"}:
        with path.open("r", encoding="utf-8-sig") as handle:
            for line_number, line in enumerate(handle, start=1):
                if not line.strip():
                    continue
                value = json.loads(line)
                if not isinstance(value, dict):
                    raise ValueError(f"{path}:{line_number} is not a JSON object")
                yield value
        return

    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(payload, dict):
        items = payload.get("items") or payload.get("data") or payload.get("results")
        if items is None:
            yield payload
            return
        payload = items
    if not isinstance(payload, list):
        raise ValueError(f"Expected a JSON list or object in {path}")
    for index, value in enumerate(payload):
        if not isinstance(value, dict):
            raise ValueError(f"{path}: item {index} is not a JSON object")
        yield value


def _iter_xlsx(path: Path) -> Iterator[Mapping[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            try:
                raw_header = next(rows)
            except StopIteration:
                continue
            headers = [str(value).strip() if value is not None else "" for value in raw_header]
            if not any(headers):
                continue
            for values in rows:
                row = {headers[index]: value for index, value in enumerate(values) if headers[index]}
                if any(value not in (None, "") for value in row.values()):
                    yield row
    finally:
        workbook.close()


def iter_rows(path: Path) -> Iterator[Mapping[str, object]]:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        yield from _iter_delimited(path, ",")
    elif suffix == ".tsv":
        yield from _iter_delimited(path, "\t")
    elif suffix in {".json", ".jsonl", ".ndjson"}:
        yield from _iter_json(path)
    elif suffix == ".xlsx":
        yield from _iter_xlsx(path)
    else:
        raise ValueError(f"Unsupported file extension: {path.suffix}")
