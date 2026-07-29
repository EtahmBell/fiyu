from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from .address_research import _prepare_result_evidence, _result_from_evidence_row
from .database import connect
from .public_catalog import ensure_public_schema
from .sqlite_snapshot import readonly_sqlite_snapshot

ADDRESS_REVIEW_COLUMNS = [
    "public_restaurant_id",
    "name_ja",
    "name_en",
    "discovery_areas",
    "candidate_evidence_id",
    "evidence_fingerprint",
    "effective_decision_fingerprint",
    "candidate_address",
    "postal_code",
    "prefecture",
    "municipality_or_ward",
    "neighborhood",
    "street_or_block",
    "building",
    "floor",
    "suite_or_unit",
    "entrance",
    "agreed_core_address",
    "core_address_verified",
    "full_address_verified",
    "component_agreement_json",
    "material_conflicting_components",
    "non_material_conflicting_components",
    "unresolved_address_detail",
    "proposed_precision",
    "proposed_approximate",
    "map_eligible_after_geocoding",
    "source_type",
    "source_url",
    "source_title",
    "identity_status",
    "identity_confidence",
    "supporting_sources_json",
    "conflicting_sources_json",
    "warnings_json",
    "current_acceptance_status",
    "current_resolution_status",
    "current_review_reasons",
    "previous_acceptance_status",
    "latest_decision_at",
    "latest_decision_source",
    "proposed_decision",
    "reviewer_decision",
    "reviewer_notes",
    "reviewed_by",
    "reviewed_at",
]

_DRY_RUN_REQUIRED_COLUMNS = {
    "public_restaurants": {
        "place_id", "name_ja", "name_en", "discovery_areas_json", "is_published",
        "map_display_eligible",
    },
    "address_evidence": {
        "id", "public_restaurant_id", "identity_status", "identity_confidence",
        "matched_name", "branch_name", "address_raw", "postal_code", "prefecture",
        "municipality_or_ward", "neighborhood", "street_or_block", "building", "floor",
        "suite_or_unit", "entrance", "component_agreement_json", "source_evidence_json",
        "conflicting_addresses_json", "search_queries_json", "warnings_json",
        "recommended_action", "research_summary", "acceptance_status",
        "acceptance_reasons_json", "evidence_fingerprint", "created_at",
    },
    "address_decision_audits": {
        "id", "address_evidence_id", "original_evidence_fingerprint",
        "component_agreement_json", "acceptance_status", "resolution_status",
        "acceptance_reasons_json", "created_at",
    },
    "address_review_decisions": {
        "id", "address_evidence_id", "evidence_fingerprint", "reviewer_decision",
        "reviewer_notes", "created_at",
    },
    "verified_restaurant_addresses": {"public_restaurant_id", "status"},
}


def _require_current_dry_run_schema(connection) -> None:
    problems: list[str] = []
    for table, required_columns in _DRY_RUN_REQUIRED_COLUMNS.items():
        columns = {
            str(row["name"])
            for row in connection.execute(f'PRAGMA table_info("{table}")').fetchall()
        }
        if not columns:
            problems.append(f"missing table {table}")
            continue
        missing = sorted(required_columns - columns)
        if missing:
            problems.append(f"{table} missing columns: {', '.join(missing)}")
    if problems:
        raise RuntimeError(
            "address-review dry-run requires the current database schema; run "
            "`python -m fiyu.public_cli --db PATH init` explicitly first. "
            + "; ".join(problems)
        )
ADDRESS_REVIEW_DECISIONS = {
    "approve", "approve_core_location", "approve_full_address", "reject", "unresolved"
}

_MANUAL_DECISION_STATUS = {
    "approve": ("review_approved", "address_verified"),
    "approve_core_location": ("review_approved", "address_verified"),
    "approve_full_address": ("review_approved", "address_verified"),
    "reject": ("rejected", "address_rejected"),
    "unresolved": ("needs_review", "address_needs_review"),
}
_ORIGINAL_RESOLUTION_STATUS = {
    "accepted": "address_verified",
    "provisional": "address_provisionally_accepted",
    "review_approved": "address_verified",
    "needs_review": "address_needs_review",
    "conflicting": "address_conflicting",
    "failed": "address_research_failed",
    "not_found": "address_not_found",
    "rejected": "address_rejected",
}
_EXCEL_DATE_PATTERN = re.compile(
    r"^(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[-\s]\d{1,2}$",
    re.IGNORECASE,
)


def _json_list(value: str | None) -> list[object]:
    try:
        parsed = json.loads(value or "[]")
    except json.JSONDecodeError:
        return []
    return parsed if isinstance(parsed, list) else []


def _json_object(value: str | None) -> dict[str, Any]:
    try:
        parsed = json.loads(value or "{}")
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _latest_effective_decision(connection, evidence) -> dict[str, Any]:
    """Resolve immutable evidence through manual-over-deterministic decision precedence."""

    evidence_id = int(evidence["id"])
    fingerprint = str(evidence["evidence_fingerprint"])
    audit = connection.execute(
        """
        SELECT * FROM address_decision_audits
        WHERE address_evidence_id=? AND original_evidence_fingerprint=?
        ORDER BY created_at DESC, id DESC LIMIT 1
        """,
        (evidence_id, fingerprint),
    ).fetchone()
    manual = connection.execute(
        """
        SELECT * FROM address_review_decisions
        WHERE address_evidence_id=? AND evidence_fingerprint=?
        ORDER BY created_at DESC, id DESC LIMIT 1
        """,
        (evidence_id, fingerprint),
    ).fetchone()

    agreement = _json_object(
        audit["component_agreement_json"] if audit else evidence["component_agreement_json"]
    )
    previous_status = str(evidence["acceptance_status"])
    if manual:
        manual_decision = str(manual["reviewer_decision"]).casefold()
        acceptance_status, resolution_status = _MANUAL_DECISION_STATUS.get(
            manual_decision, ("needs_review", "address_needs_review")
        )
        reasons = [f"manual_review:{manual_decision}"]
        if manual["reviewer_notes"]:
            reasons.append(str(manual["reviewer_notes"]))
        decision_source = "manual_review"
        decision_id = int(manual["id"])
        decision_at = str(manual["created_at"])
    elif audit:
        acceptance_status = str(audit["acceptance_status"])
        resolution_status = str(audit["resolution_status"])
        reasons = _json_list(audit["acceptance_reasons_json"])
        decision_source = "deterministic_recalculation"
        decision_id = int(audit["id"])
        decision_at = str(audit["created_at"])
    else:
        acceptance_status = previous_status
        resolution_status = _ORIGINAL_RESOLUTION_STATUS.get(
            acceptance_status, "address_needs_review"
        )
        reasons = _json_list(evidence["acceptance_reasons_json"])
        decision_source = "original_research"
        decision_id = evidence_id
        decision_at = str(evidence["created_at"])

    decision_payload = {
        "address_evidence_id": evidence_id,
        "evidence_fingerprint": fingerprint,
        "decision_source": decision_source,
        "decision_id": decision_id,
        "decision_at": decision_at,
        "acceptance_status": acceptance_status,
        "resolution_status": resolution_status,
        "review_reasons": reasons,
        "component_agreement": agreement,
    }
    encoded = json.dumps(
        decision_payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return {
        **decision_payload,
        "effective_decision_fingerprint": hashlib.sha256(encoded).hexdigest(),
    }


def _latest_evidence_rows(connection, *, limit: int | None = None):
    query = """
        SELECT p.place_id, p.name_ja, p.name_en, p.discovery_areas_json,
               p.is_published, p.map_display_eligible, e.*
        FROM address_evidence e
        JOIN public_restaurants p ON p.place_id=e.public_restaurant_id
        WHERE e.id=(
            SELECT MAX(newest.id) FROM address_evidence newest
            WHERE newest.public_restaurant_id=e.public_restaurant_id
        )
        ORDER BY p.fiyu_score DESC, p.place_id
    """
    rows = connection.execute(query).fetchall()
    return rows if limit is None else rows[:limit]


def _effective_export_row(connection, evidence) -> dict[str, object]:
    decision = _latest_effective_decision(connection, evidence)
    agreement = decision["component_agreement"]
    result = _prepare_result_evidence(_result_from_evidence_row(evidence))
    sources = [source.model_dump(mode="json") for source in result.source_evidence]
    conflicts = [
        conflict.model_dump(mode="json")
        for conflict in result.conflicting_address_candidates
    ]
    primary = next(
        (
            source
            for source in sources
            if source.get("supports_candidate_address")
            and source.get("address_temporality") not in {"historical", "future"}
        ),
        sources[0] if sources else {},
    )
    core_verified = bool(agreement.get("core_address_verified"))
    full_verified = bool(agreement.get("full_address_verified"))
    return {
        "public_restaurant_id": evidence["place_id"],
        "name_ja": evidence["name_ja"] or "",
        "name_en": evidence["name_en"] or "",
        "discovery_areas": evidence["discovery_areas_json"],
        "candidate_evidence_id": evidence["id"],
        "evidence_fingerprint": evidence["evidence_fingerprint"],
        "effective_decision_fingerprint": decision["effective_decision_fingerprint"],
        "candidate_address": result.address_raw or "",
        "postal_code": result.postal_code or "",
        "prefecture": result.prefecture or "",
        "municipality_or_ward": result.municipality_or_ward or "",
        "neighborhood": result.neighborhood or "",
        "street_or_block": result.street_or_block or "",
        "building": result.building or "",
        "floor": result.floor or "",
        "suite_or_unit": result.suite_or_unit or "",
        "entrance": result.entrance or "",
        "agreed_core_address": agreement.get("agreed_core_address") or "",
        "core_address_verified": core_verified,
        "full_address_verified": full_verified,
        "component_agreement_json": json.dumps(agreement, ensure_ascii=False, sort_keys=True),
        "material_conflicting_components": json.dumps(
            agreement.get("material_conflicting_components", []), ensure_ascii=False
        ),
        "non_material_conflicting_components": json.dumps(
            agreement.get("non_material_conflicting_components", []), ensure_ascii=False
        ),
        "unresolved_address_detail": agreement.get("unresolved_address_detail") or "",
        "proposed_precision": agreement.get("proposed_location_precision") or "unknown",
        "proposed_approximate": bool(agreement.get("map_location_approximate")),
        "map_eligible_after_geocoding": core_verified,
        "source_type": primary.get("source_type", ""),
        "source_url": primary.get("source_url", ""),
        "source_title": primary.get("source_title", ""),
        "identity_status": result.identity_status,
        "identity_confidence": result.identity_confidence,
        "supporting_sources_json": json.dumps(sources, ensure_ascii=False, sort_keys=True),
        "conflicting_sources_json": json.dumps(conflicts, ensure_ascii=False, sort_keys=True),
        "warnings_json": json.dumps(result.warnings, ensure_ascii=False),
        "current_acceptance_status": decision["acceptance_status"],
        "current_resolution_status": decision["resolution_status"],
        "current_review_reasons": json.dumps(decision["review_reasons"], ensure_ascii=False),
        "previous_acceptance_status": evidence["acceptance_status"],
        "latest_decision_at": decision["decision_at"],
        "latest_decision_source": decision["decision_source"],
        "proposed_decision": (
            "approve_full_address" if full_verified
            else "approve_core_location" if core_verified
            else "unresolved"
        ),
        "reviewer_decision": "",
        "reviewer_notes": "",
        "reviewed_by": "",
        "reviewed_at": "",
    }


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _write_review_file(output: Path, rows: list[dict[str, object]]) -> None:
    if output.suffix.casefold() == ".xlsx":
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "address_review"
        sheet.append(ADDRESS_REVIEW_COLUMNS)
        for row in rows:
            sheet.append([_cell_text(row.get(column)) for column in ADDRESS_REVIEW_COLUMNS])
        for row in sheet.iter_rows():
            for cell in row:
                cell.number_format = "@"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(output)
        return
    if output.suffix.casefold() != ".csv":
        raise ValueError("address review output must use a .csv or .xlsx extension")
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=ADDRESS_REVIEW_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _read_review_file(source_path: Path) -> list[dict[str, str]]:
    if source_path.suffix.casefold() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(source_path, read_only=True, data_only=False)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [_cell_text(value) for value in next(values, ())]
        return [
            {header: _cell_text(value) for header, value in zip(headers, row, strict=False)}
            for row in values
            if any(value is not None and value != "" for value in row)
        ]
    if source_path.suffix.casefold() != ".csv":
        raise ValueError("address review input must use a .csv or .xlsx extension")
    with source_path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def export_address_review(
    db_path: str | Path,
    output_path: str | Path,
    *,
    limit: int = 100,
) -> int:
    ensure_public_schema(db_path)
    with connect(db_path) as connection:
        export_rows: list[dict[str, object]] = []
        for evidence in _latest_evidence_rows(connection):
            if not evidence["is_published"] or evidence["map_display_eligible"]:
                continue
            row = _effective_export_row(connection, evidence)
            if row["current_resolution_status"] != "address_needs_review":
                continue
            export_rows.append(row)
            if len(export_rows) >= limit:
                break
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    _write_review_file(output, export_rows)
    return len(export_rows)


def import_address_review(
    db_path: str | Path,
    input_path: str | Path,
    *,
    dry_run: bool = False,
) -> dict[str, object]:
    if not dry_run:
        ensure_public_schema(db_path)
    source_path = Path(input_path)
    rows = _read_review_file(source_path)
    reports: list[dict[str, object]] = []
    accepted: list[tuple[dict[str, str], object, str, dict[str, object]]] = []
    seen_restaurants: set[str] = set()
    if dry_run:
        database = readonly_sqlite_snapshot(db_path)
    else:
        database = connect(db_path)
    with database as connection:
        if dry_run:
            _require_current_dry_run_schema(connection)
        for line, raw in enumerate(rows, start=2):
            place_id = (raw.get("public_restaurant_id") or "").strip()
            decision = (raw.get("reviewer_decision") or "").strip().casefold()
            errors: list[str] = []
            if not decision:
                reports.append(
                    {
                        "line": line,
                        "place_id": place_id or None,
                        "valid": True,
                        "skipped": True,
                        "errors": [],
                    }
                )
                continue
            if place_id in seen_restaurants:
                errors.append("duplicate restaurant decision")
            seen_restaurants.add(place_id)
            if decision not in ADDRESS_REVIEW_DECISIONS:
                errors.append(
                    "reviewer_decision must be approve_core_location, approve_full_address, reject, or unresolved"
                )
            try:
                evidence_id = int(raw.get("candidate_evidence_id") or "")
            except ValueError:
                evidence_id = -1
                errors.append("invalid candidate_evidence_id")
            evidence = connection.execute(
                """
                SELECT e.*, p.place_id, p.name_ja, p.name_en, p.discovery_areas_json,
                       p.discovery_areas_json AS stored_discovery_areas,
                       p.is_published, p.map_display_eligible,
                       EXISTS(SELECT 1 FROM verified_restaurant_addresses v
                              WHERE v.public_restaurant_id=p.place_id
                                AND v.status IN ('address_verified', 'location_verified'))
                           AS has_verified_address
                FROM address_evidence e
                JOIN public_restaurants p ON p.place_id=e.public_restaurant_id
                WHERE e.id=? AND e.public_restaurant_id=?
                """,
                (evidence_id, place_id),
            ).fetchone()
            if evidence is None:
                errors.append("missing persisted address evidence")
            else:
                latest_evidence_id = connection.execute(
                    "SELECT MAX(id) FROM address_evidence WHERE public_restaurant_id=?",
                    (place_id,),
                ).fetchone()[0]
                if evidence_id != latest_evidence_id:
                    errors.append("candidate evidence has been superseded by newer evidence")
                effective = _effective_export_row(connection, evidence)
                immutable = {
                    field: _cell_text(effective[field])
                    for field in (
                        "name_ja", "name_en", "discovery_areas", "evidence_fingerprint",
                        "effective_decision_fingerprint", "candidate_address", "postal_code",
                        "prefecture", "municipality_or_ward", "neighborhood",
                        "street_or_block", "building", "floor", "suite_or_unit", "entrance",
                        "agreed_core_address", "component_agreement_json",
                    )
                }
                for field, expected in immutable.items():
                    if (raw.get(field) or "").strip() != expected:
                        errors.append(f"{field} does not match persisted evidence")
                supplied_street = (raw.get("street_or_block") or "").strip()
                if supplied_street != immutable["street_or_block"] and (
                    _EXCEL_DATE_PATTERN.fullmatch(supplied_street)
                    or re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:[ T].*)?", supplied_street)
                ):
                    errors.append(
                        "street_or_block appears converted to a spreadsheet date; "
                        "re-export and keep address columns formatted as text"
                    )
                if effective["current_resolution_status"] != "address_needs_review":
                    errors.append(
                        "record no longer has current effective status address_needs_review"
                    )
                if not evidence["is_published"]:
                    errors.append("restaurant is not published")
                if evidence["map_display_eligible"]:
                    errors.append("restaurant already has a verified map location")
                if decision in {"approve", "approve_full_address"} and not effective[
                    "full_address_verified"
                ]:
                    errors.append("full-address approval requires full_address_verified evidence")
                if decision == "approve_core_location" and not effective[
                    "core_address_verified"
                ]:
                    errors.append("core-location approval requires an agreed verified core")
                if decision.startswith("approve") and not effective["candidate_address"]:
                    errors.append("approved evidence requires a candidate address")
                if decision.startswith("approve") and evidence["has_verified_address"]:
                    errors.append("restaurant already has a verified address")
            reviewed_by = (raw.get("reviewed_by") or "").strip()
            reviewed_at = (raw.get("reviewed_at") or "").strip()
            if not reviewed_by or not reviewed_at:
                errors.append("review decisions require reviewed_by and reviewed_at")
            if reviewed_at:
                try:
                    parsed_date = date.fromisoformat(reviewed_at)
                    if parsed_date.isoformat() != reviewed_at:
                        raise ValueError
                except ValueError:
                    errors.append("reviewed_at must be YYYY-MM-DD")
            reports.append(
                {
                    "line": line,
                    "place_id": place_id or None,
                    "valid": not errors,
                    "errors": errors,
                }
            )
            if not errors and evidence is not None:
                accepted.append((raw, evidence, decision, effective))

        failures = sum(not report["valid"] for report in reports)
        if not failures and not dry_run:
            now = datetime.now(UTC).isoformat()
            for raw, evidence, decision, effective in accepted:
                place_id = str(evidence["public_restaurant_id"])
                connection.execute(
                    """
                    INSERT INTO address_review_decisions (
                        public_restaurant_id, address_evidence_id, reviewer_decision,
                        reviewer_notes, reviewed_by, reviewed_at, import_provenance,
                        evidence_fingerprint, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        place_id,
                        evidence["id"],
                        decision,
                        (raw.get("reviewer_notes") or "").strip() or None,
                        raw["reviewed_by"].strip(),
                        raw["reviewed_at"].strip(),
                        str(source_path.resolve()),
                        evidence["evidence_fingerprint"],
                        now,
                    ),
                )
                if decision in {"approve", "approve_core_location", "approve_full_address"}:
                    approve_full = decision in {"approve", "approve_full_address"}
                    selected_address = (
                        effective["candidate_address"]
                        if approve_full
                        else effective["agreed_core_address"]
                    )
                    sources = _json_list(str(effective["supporting_sources_json"]))
                    references = [
                        source.get("source_url")
                        for source in sources
                        if isinstance(source, dict)
                        and source.get("source_url")
                        and source.get("address_temporality") not in {"historical", "future"}
                    ]
                    connection.execute(
                        """
                        INSERT INTO verified_restaurant_addresses (
                            public_restaurant_id, address_evidence_id, address_raw, postal_code,
                            prefecture, municipality_or_ward, neighborhood, street_or_block,
                            building, floor, suite_or_unit, entrance, verified_core_address,
                            geocoding_address, core_address_verified, full_address_verified,
                            unresolved_address_detail, approved_location_precision,
                            map_location_approximate,
                            verification_method, evidence_references_json, verified_by,
                            verified_at, status, address_confidence_tier,
                            decision_fingerprint, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?,
                                  'manual_address_review', ?, ?, ?, 'address_verified',
                                  'manual', ?, ?, ?)
                        ON CONFLICT(public_restaurant_id) DO NOTHING
                        """,
                        (
                            place_id,
                            evidence["id"],
                            selected_address,
                            effective["postal_code"],
                            effective["prefecture"],
                            effective["municipality_or_ward"],
                            effective["neighborhood"],
                            effective["street_or_block"],
                            effective["building"] if approve_full else None,
                            effective["floor"] if approve_full else None,
                            effective["suite_or_unit"] if approve_full else None,
                            effective["entrance"] if approve_full else None,
                            effective["agreed_core_address"],
                            effective["agreed_core_address"],
                            int(approve_full),
                            effective["unresolved_address_detail"],
                            effective["proposed_precision"],
                            int(not approve_full or bool(effective["proposed_approximate"])),
                            json.dumps(references, ensure_ascii=False),
                            raw["reviewed_by"].strip(),
                            raw["reviewed_at"].strip(),
                            effective["effective_decision_fingerprint"],
                            now,
                            now,
                        ),
                    )
                    status = "address_verified"
                elif decision == "reject":
                    status = "address_rejected"
                else:
                    status = "address_needs_review"
                connection.execute(
                    "UPDATE public_restaurants SET address_resolution_status=?, updated_at=? WHERE place_id=?",
                    (status, now, place_id),
                )
                if decision in {"approve", "approve_core_location", "approve_full_address"}:
                    connection.execute(
                        """
                        UPDATE public_restaurants SET verified_core_address=?,
                            core_address_verified=1, full_address_verified=?,
                            unresolved_address_detail=?, map_location_approximate=?, updated_at=?
                        WHERE place_id=?
                        """,
                        (effective["agreed_core_address"], int(approve_full),
                         effective["unresolved_address_detail"],
                         int(not approve_full or bool(effective["proposed_approximate"])),
                         now, place_id),
                    )
            connection.commit()
    return {
        "rows": len(reports),
        "valid": len(accepted),
        "validation_failures": failures,
        "updated": len(accepted) if not failures and not dry_run else 0,
        "dry_run": dry_run,
        "reports": reports,
    }
